from __future__ import annotations

import asyncio
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]
from sqlalchemy import delete, func, insert, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import SessionLocal
from app.core.errors import AppError
from app.models import HuaXingInventory
from app.schemas import HuaXingInventoryRead
from app.services.common import contains_any

EXPECTED_HEADERS = (
    "首次入库日期",
    "仓库",
    "货品编码",
    "货品名称",
    "型号",
    "数量",
    "单位",
    "申购人",
    "申购部门",
    "子项号名称",
)
# 用户口径「首次入库时间」对应报表实际表头「首次入库日期」，匹配前归一化。
HEADER_ALIASES = {"首次入库时间": "首次入库日期"}
MAX_IMPORT_BYTES = 50 * 1024 * 1024
INSERT_BATCH_SIZE = 2_000
# 全字段相同视为重复行：上游报表常含重复行，逐字导成一条即可（保留行号提示）。
DEDUPE_FIELDS = (
    "first_inbound_date",
    "warehouse",
    "material_code",
    "name",
    "model_spec",
    "quantity",
    "unit_name",
    "purchaser",
    "purchase_department",
    "subitem_no_name",
)


def _dedupe_key(row: dict[str, object]) -> tuple[object, ...]:
    return tuple(row[field] for field in DEDUPE_FIELDS)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _cell_quantity(value: object, *, row_number: int) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise AppError(
            "HUAXING_IMPORT_INVALID_QUANTITY",
            f"第 {row_number} 行数量不是有效数值",
            details={"row": row_number},
        ) from exc


def _validate_length(value: str, maximum: int, *, row_number: int, header: str) -> None:
    if len(value) > maximum:
        raise AppError(
            "HUAXING_IMPORT_VALUE_TOO_LONG",
            f"第 {row_number} 行“{header}”超过 {maximum} 个字符",
            details={"row": row_number, "column": header, "max_length": maximum},
        )


def parse_huaxing_workbook_file(path: Path) -> list[dict[str, object]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise AppError("INVALID_EXCEL_FILE", "无法读取 Excel 文件，请确认文件格式正确") from exc
    except Exception as exc:
        raise AppError(
            "INVALID_EXCEL_FILE",
            "读取 Excel 文件时发生未知错误，请确认文件未被损坏",
        ) from exc

    try:
        worksheet = workbook.active
        header_row_number = 0
        header_indexes: dict[str, int] = {}
        for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20), start=1):
            raw_headers = [_cell_text(cell.value) for cell in row]
            headers = [HEADER_ALIASES.get(header, header) for header in raw_headers]
            indexes = {header: index for index, header in enumerate(headers) if header}
            if all(header in indexes for header in EXPECTED_HEADERS):
                header_row_number = row_number
                header_indexes = {header: indexes[header] for header in EXPECTED_HEADERS}
                break
        if not header_indexes:
            raise AppError(
                "HUAXING_IMPORT_HEADERS_MISSING",
                "Excel 缺少必需列：首次入库日期、仓库、货品编码、货品名称、型号、数量、"
                "单位、申购人、申购部门、子项号名称",
            )

        date_index = header_indexes["首次入库日期"]
        quantity_index = header_indexes["数量"]
        rows: list[dict[str, object]] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row_number + 1), start=header_row_number + 1
        ):
            values = {
                header: _cell_text(row[index].value) if index < len(row) else ""
                for header, index in header_indexes.items()
            }
            if not any(values.values()):
                continue
            material_code = values["货品编码"]
            if not material_code:
                raise AppError(
                    "HUAXING_IMPORT_CODE_REQUIRED",
                    f"第 {row_number} 行缺少货品编码",
                    details={"row": row_number},
                )
            _validate_length(material_code, 64, row_number=row_number, header="货品编码")
            _validate_length(values["仓库"], 128, row_number=row_number, header="仓库")
            _validate_length(values["货品名称"], 255, row_number=row_number, header="货品名称")
            _validate_length(values["型号"], 255, row_number=row_number, header="型号")
            _validate_length(values["单位"], 32, row_number=row_number, header="单位")
            _validate_length(values["申购人"], 128, row_number=row_number, header="申购人")
            _validate_length(values["申购部门"], 128, row_number=row_number, header="申购部门")
            _validate_length(values["子项号名称"], 255, row_number=row_number, header="子项号名称")
            rows.append(
                {
                    "first_inbound_date": (
                        _cell_date(row[date_index].value) if date_index < len(row) else None
                    ),
                    "warehouse": values["仓库"] or None,
                    "material_code": material_code,
                    "name": values["货品名称"] or None,
                    "model_spec": values["型号"] or None,
                    "quantity": (
                        _cell_quantity(row[quantity_index].value, row_number=row_number)
                        if quantity_index < len(row)
                        else None
                    ),
                    "unit_name": values["单位"] or None,
                    "purchaser": values["申购人"] or None,
                    "purchase_department": values["申购部门"] or None,
                    "subitem_no_name": values["子项号名称"] or None,
                }
            )
        if not rows:
            raise AppError("HUAXING_IMPORT_EMPTY", "Excel 中没有可导入的华星库存数据")
        return rows
    finally:
        workbook.close()


async def process_import_file(file_path: Path) -> dict[str, object]:
    """异步导入处理器：解析（线程池）→ 全量替换 → 返回导入条数与去重统计。"""
    rows = await asyncio.to_thread(parse_huaxing_workbook_file, file_path)
    return await _replace_rows(rows)


async def _replace_rows(rows: list[dict[str, object]]) -> dict[str, object]:
    """全量替换华星库存表；完全重复的行只保留一条（上游报表常重复导出）。"""
    seen: set[tuple[object, ...]] = set()
    deduplicated: list[dict[str, object]] = []
    for row in rows:
        key = _dedupe_key(row)
        if key in seen:
            continue
        seen.add(key)
        deduplicated.append(row)
    async with SessionLocal() as session:
        await session.execute(delete(HuaXingInventory))
        for offset in range(0, len(deduplicated), INSERT_BATCH_SIZE):
            batch = deduplicated[offset : offset + INSERT_BATCH_SIZE]
            await session.execute(insert(HuaXingInventory), batch)
        await session.commit()
    return {
        "imported_count": len(deduplicated),
        "deduplicated_count": len(rows) - len(deduplicated),
    }


async def search_huaxing_inventory(
    session: AsyncSession,
    *,
    keyword: str | None,
    warehouse: str | None,
    purchase_department: str | None,
    purchaser: str | None,
    page: int,
    page_size: int,
) -> tuple[list[HuaXingInventoryRead], int]:
    query = select(HuaXingInventory)
    for columns, value in (
        (
            (
                HuaXingInventory.material_code,
                HuaXingInventory.name,
                HuaXingInventory.model_spec,
                HuaXingInventory.purchaser,
            ),
            keyword,
        ),
        ((HuaXingInventory.warehouse,), warehouse),
        ((HuaXingInventory.purchase_department,), purchase_department),
        ((HuaXingInventory.purchaser,), purchaser),
    ):
        condition = contains_any(columns, value)
        if condition is not None:
            query = query.where(condition)
    total = int((await session.scalar(select(func.count()).select_from(query.subquery()))) or 0)
    result = await session.scalars(
        query.order_by(HuaXingInventory.id).offset((page - 1) * page_size).limit(page_size)
    )
    items = [
        HuaXingInventoryRead(
            id=item.id,
            first_inbound_date=item.first_inbound_date,
            warehouse=item.warehouse,
            material_code=item.material_code,
            name=item.name,
            model_spec=item.model_spec,
            quantity=item.quantity,
            unit_name=item.unit_name,
            purchaser=item.purchaser,
            purchase_department=item.purchase_department,
            subitem_no_name=item.subitem_no_name,
        )
        for item in result.all()
    ]
    return items, total
