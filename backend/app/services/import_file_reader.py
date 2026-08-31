"""表格导入文件读取：按扩展名统一读取 xlsx/xlsm/xls/csv 为二维单元格值。

物料编码库与华星库存共用底层读取器。这里只负责把文件解析成
``list[list[object]]``（每行是原始单元格值列表），表头定位、字段校验、
日期/数量换算等业务逻辑仍由各 service 负责，避免把不同报表的表头与校验
规则耦合进底层读取器。

约定：
- xlsx/xlsm 用 openpyxl（只读模式，读取缓存值）。
- xls 用 xlrd（旧版 Excel 二进制格式，openpyxl 不支持）。
- csv 用标准库 csv（自动探测 UTF-8(BOM)/GB18030 编码，适配中文 Excel 导出）。
"""

from __future__ import annotations

import csv
from io import StringIO
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from app.core.errors import AppError

# 与 API 层校验保持一致；改动时两处都要同步。
SUPPORTED_IMPORT_SUFFIXES = (".xlsx", ".xlsm", ".xls", ".csv")

_EXCEL_SUFFIXES = (".xlsx", ".xlsm")
_CSV_ENCODINGS = ("utf-8-sig", "gb18030")


def read_tabular_rows(path: Path) -> list[list[object]]:
    """读取表格文件为二维原始值；不支持的类型抛 UNSUPPORTED_EXCEL_FILE。"""
    suffix = path.suffix.lower()
    if suffix in _EXCEL_SUFFIXES:
        return _read_xlsx(path)
    if suffix == ".xls":
        return _read_xls(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise AppError("UNSUPPORTED_EXCEL_FILE", "仅支持 .xls、.xlsx 或 .csv 格式的表格文件")


def _read_xlsx(path: Path) -> list[list[object]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise AppError("INVALID_EXCEL_FILE", "无法读取 Excel 文件，请确认文件格式正确") from exc
    except Exception as exc:
        raise AppError(
            "INVALID_EXCEL_FILE",
            "读取 Excel 文件时发生未知错误，请确认文件未被损坏",
        ) from exc
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_xls(path: Path) -> list[list[object]]:
    try:
        import xlrd  # type: ignore[import-untyped]
    except ImportError as exc:
        raise AppError(
            "XLS_SUPPORT_UNAVAILABLE",
            "当前环境未安装 .xls 解析依赖，请联系管理员",
        ) from exc
    try:
        book = xlrd.open_workbook(str(path))
    except xlrd.biffh.XLRDError as exc:
        raise AppError("INVALID_EXCEL_FILE", "无法读取 Excel 文件，请确认文件格式正确") from exc

    sheet = book.sheet_by_index(0)
    datemode = book.datemode
    rows: list[list[object]] = []
    for row_index in range(sheet.nrows):
        values = sheet.row_values(row_index)
        types = sheet.row_types(row_index)
        normalized: list[object] = []
        for value, cell_type in zip(values, types, strict=True):
            if cell_type == xlrd.XL_CELL_DATE:
                normalized.append(xlrd.xldate.xldate_as_datetime(value, datemode))
            elif cell_type == xlrd.XL_CELL_EMPTY:
                normalized.append(None)
            else:
                normalized.append(value)
        rows.append(normalized)
    return rows


def _read_csv(path: Path) -> list[list[object]]:
    content: str | None = None
    last_error: UnicodeDecodeError | None = None
    for encoding in _CSV_ENCODINGS:
        try:
            content = path.read_bytes().decode(encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    if content is None:
        raise AppError(
            "INVALID_CSV_FILE",
            "无法解析 CSV 文件编码，请使用 UTF-8 或 GBK 编码",
        ) from last_error
    reader = csv.reader(StringIO(content, newline=""))
    return [list(row) for row in reader]
