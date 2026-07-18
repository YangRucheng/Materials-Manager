from __future__ import annotations

import json
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.formatting.rule import FormulaRule  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.datavalidation import (  # type: ignore[import-untyped]
    DataValidation,
)

from app.core.config import settings

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _load_spec(file_name: str) -> dict[str, Any]:
    path = settings.template_dir / file_name
    return json.loads(path.read_text(encoding="utf-8"))  # type: ignore[no-any-return]


def _style(cell: Any, style: dict[str, Any]) -> None:
    color = style.get("fill")
    cell.font = Font(
        name=style.get("font_name", "等线"),
        size=style.get("font_size", 11),
        bold=style.get("bold", False),
    )
    cell.alignment = Alignment(
        horizontal=style.get("horizontal"),
        vertical=style.get("vertical"),
        wrap_text=style.get("wrap_text", False),
    )
    if color:
        cell.fill = PatternFill("solid", fgColor=f"FF{color}")
    if style.get("border"):
        side = Side(style="thin", color="FF000000")
        cell.border = Border(left=side, right=side, top=side, bottom=side)
    if number_format := style.get("number_format"):
        cell.number_format = number_format


def _set_dimensions(sheet: Any, spec: dict[str, Any]) -> None:
    for column, width in spec.get("column_widths", {}).items():
        sheet.column_dimensions[column].width = width
    for row, height in spec.get("row_heights", {}).items():
        sheet.row_dimensions[int(row)].height = height


def _cell_value(row: dict[str, Any], column: dict[str, Any]) -> Any:
    field = column.get("field")
    value = row.get(field) if field else None
    result = value if value not in (None, "") else column.get("default")
    if isinstance(result, str) and result.startswith(("=", "+", "-", "@")):
        return f"'{result}"
    return result


def _material_code_workbook(spec: dict[str, Any], rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec["sheet"]["title"]
    sheet.sheet_view.showGridLines = spec["sheet"].get("show_grid_lines", True)
    _set_dimensions(sheet, spec["sheet"])
    styles = spec["styles"]

    for merged_range in spec["sheet"].get("merged_cells", []):
        sheet.merge_cells(merged_range)
    title = sheet[spec["sheet"]["title_cell"]]
    title.value = spec["sheet"]["title"]
    _style(title, styles["title"])
    for section in spec["sheet"].get("sections", []):
        cell = sheet[section["cell"]]
        cell.value = section["label"]
        _style(cell, styles["section"])

    header_row = spec["sheet"]["header_row"]
    required_row = spec["sheet"]["required_row"]
    instruction_row = spec["sheet"]["instruction_row"]
    sheet[f"B{instruction_row}"] = "说明"
    _style(sheet[f"B{instruction_row}"], styles["instruction"])
    for column in spec["columns"]:
        letter = column["column"]
        for row_no, key in (
            (header_row, "header"),
            (required_row, "required"),
            (instruction_row, "instruction"),
        ):
            cell = sheet[f"{letter}{row_no}"]
            cell.value = column.get(key)
            _style(cell, styles[column.get(f"{key}_style", "instruction")])

    data_start = spec["sheet"]["data_start_row"]
    data_row_count = max(len(rows), spec["sheet"].get("minimum_data_rows", 1))
    for offset in range(data_row_count):
        row = rows[offset] if offset < len(rows) else {}
        for column in spec["columns"]:
            cell = sheet[f"{column['column']}{data_start + offset}"]
            cell.value = _cell_value(row, column)
            _style(cell, styles[column.get("data_style", "data")])
        sheet.row_dimensions[data_start + offset].height = 24
    sheet.print_area = f"B2:AA{data_start + data_row_count - 1}"
    sheet.freeze_panes = f"B{data_start}"
    return workbook


def _purchase_application_workbook(
    spec: dict[str, Any], rows: list[dict[str, Any]]
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec["sheet"]["title"]
    sheet.freeze_panes = spec["sheet"].get("freeze_panes")
    _set_dimensions(sheet, spec["sheet"])
    styles = spec["styles"]
    header_row = spec["sheet"]["header_row"]
    data_start = spec["sheet"]["data_start_row"]

    for column in spec["columns"]:
        header = sheet[f"{column['column']}{header_row}"]
        header.value = column["header"]
        _style(header, styles["header"])
    sheet.row_dimensions[header_row].height = 34

    for offset, row in enumerate(rows):
        row_no = data_start + offset
        for column in spec["columns"]:
            cell = sheet[f"{column['column']}{row_no}"]
            cell.value = _cell_value(row, column)
            _style(cell, styles[column.get("style", "data")])
        sheet.row_dimensions[row_no].height = 24

    last_row = max(data_start, data_start + len(rows) - 1)
    for column in spec["columns"]:
        validation_name = column.get("validation")
        if not validation_name:
            continue
        options = spec["validation_lists"][validation_name]
        validation = DataValidation(
            type="list",
            formula1='"' + ",".join(options) + '"',
            allow_blank=True,
        )
        validation.error = "请选择下拉列表中的值"
        validation.errorTitle = "无效选项"
        sheet.add_data_validation(validation)
        validation.add(f"{column['column']}{data_start}:{column['column']}{last_row}")

    if rows:
        red_fill = PatternFill("solid", fgColor="FFFCE8E6")
        sheet.conditional_formatting.add(
            f"A{data_start}:K{last_row}",
            FormulaRule(formula=[f'LEN($A{data_start})=0'], fill=red_fill),
        )
    sheet.auto_filter.ref = f"A{header_row}:K{last_row}"
    sheet.print_area = f"A{header_row}:K{last_row}"
    return workbook


def render_excel(template_file: str, rows: list[dict[str, Any]]) -> tuple[bytes, str]:
    spec = _load_spec(template_file)
    if template_file == "material-code-application.json":
        workbook = _material_code_workbook(spec, rows)
    else:
        workbook = _purchase_application_workbook(spec, rows)
    output = BytesIO()
    workbook.save(output)
    filename = spec["output_filename"].format(date=date.today().isoformat())
    return output.getvalue(), filename
