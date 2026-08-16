from __future__ import annotations

import asyncio
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]
from sqlalchemy import delete, func, insert, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import SessionLocal
from app.core.errors import AppError
from app.models import MaterialCodeLibrary
from app.schemas import MaterialCodeLibraryRead
from app.services.common import contains_any

EXPECTED_HEADERS = ("编码", "名称", "型号", "记账单位名称")
MAX_IMPORT_BYTES = 50 * 1024 * 1024
INSERT_BATCH_SIZE = 2_000


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _validate_length(value: str, maximum: int, *, row_number: int, header: str) -> None:
    if len(value) > maximum:
        raise AppError(
            "MATERIAL_CODE_IMPORT_VALUE_TOO_LONG",
            f"第 {row_number} 行“{header}”超过 {maximum} 个字符",
            details={"row": row_number, "column": header, "max_length": maximum},
        )


def _load_workbook(source: Path | bytes) -> Workbook:
    try:
        if isinstance(source, Path):
            return load_workbook(source, read_only=True, data_only=True)
        return load_workbook(BytesIO(source), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise AppError("INVALID_EXCEL_FILE", "无法读取 Excel 文件，请确认文件格式正确") from exc
    except Exception as exc:
        raise AppError(
            "INVALID_EXCEL_FILE",
            "读取 Excel 文件时发生未知错误，请确认文件未被损坏",
        ) from exc


def _parse_material_codes(worksheet: Worksheet) -> list[dict[str, str | None]]:
    header_row_number = 0
    header_indexes: dict[str, int] = {}
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20), start=1):
        headers = [_cell_text(cell.value) for cell in row]
        indexes = {header: index for index, header in enumerate(headers) if header}
        if all(header in indexes for header in EXPECTED_HEADERS):
            header_row_number = row_number
            header_indexes = {header: indexes[header] for header in EXPECTED_HEADERS}
            break
    if not header_indexes:
        raise AppError(
            "MATERIAL_CODE_IMPORT_HEADERS_MISSING",
            "Excel 缺少必需列：编码、名称、型号、记账单位名称",
        )

    rows: list[dict[str, str | None]] = []
    seen_codes: dict[str, int] = {}
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row_number + 1), start=header_row_number + 1
    ):
        values = {
            header: _cell_text(row[index].value) if index < len(row) else ""
            for header, index in header_indexes.items()
        }
        if not any(values.values()):
            continue
        material_code = values["编码"]
        if not material_code:
            raise AppError(
                "MATERIAL_CODE_IMPORT_CODE_REQUIRED",
                f"第 {row_number} 行缺少编码",
                details={"row": row_number},
            )
        if not values["记账单位名称"]:
            raise AppError(
                "MATERIAL_CODE_IMPORT_UNIT_REQUIRED",
                f"第 {row_number} 行缺少记账单位名称",
                details={"row": row_number},
            )
        if material_code in seen_codes:
            raise AppError(
                "MATERIAL_CODE_IMPORT_DUPLICATE",
                (
                    f"编码“{material_code}”在第 {seen_codes[material_code]} 行"
                    f"和第 {row_number} 行重复"
                ),
                details={
                    "material_code": material_code,
                    "first_row": seen_codes[material_code],
                    "duplicate_row": row_number,
                },
            )
        _validate_length(material_code, 64, row_number=row_number, header="编码")
        _validate_length(values["名称"], 128, row_number=row_number, header="名称")
        _validate_length(values["型号"], 255, row_number=row_number, header="型号")
        _validate_length(
            values["记账单位名称"], 32, row_number=row_number, header="记账单位名称"
        )
        seen_codes[material_code] = row_number
        rows.append(
            {
                "material_code": material_code,
                "name": values["名称"] or None,
                "model_spec": values["型号"] or None,
                "unit_name": values["记账单位名称"],
            }
        )
    if not rows:
        raise AppError("MATERIAL_CODE_IMPORT_EMPTY", "Excel 中没有可导入的物料编码数据")
    return rows


def parse_material_code_workbook(content: bytes) -> list[dict[str, str | None]]:
    if not content:
        raise AppError("EMPTY_EXCEL_FILE", "请选择需要导入的 Excel 文件")
    if len(content) > MAX_IMPORT_BYTES:
        raise AppError("EXCEL_FILE_TOO_LARGE", "Excel 文件不能超过 50 MB")
    workbook = _load_workbook(content)
    try:
        return _parse_material_codes(workbook.active)
    finally:
        workbook.close()


def parse_material_code_workbook_file(path: Path) -> list[dict[str, str | None]]:
    workbook = _load_workbook(path)
    try:
        return _parse_material_codes(workbook.active)
    finally:
        workbook.close()


async def process_import_file(file_path: Path) -> dict[str, object]:
    """异步导入处理器：解析（线程池）→ 全量替换 → 返回结果摘要。

    解析发生在任何变更之前，坏文件不会动到现有数据。
    """
    rows = await asyncio.to_thread(parse_material_code_workbook_file, file_path)
    async with SessionLocal() as session:
        await session.execute(delete(MaterialCodeLibrary))
        for offset in range(0, len(rows), INSERT_BATCH_SIZE):
            batch = rows[offset : offset + INSERT_BATCH_SIZE]
            await session.execute(insert(MaterialCodeLibrary), batch)
        await session.commit()
    return {
        "imported_count": len(rows),
        "blank_name_count": sum(row["name"] is None for row in rows),
        "blank_model_spec_count": sum(row["model_spec"] is None for row in rows),
    }


async def material_code_exists(session: AsyncSession, material_code: str) -> bool:
    """编码是否已收录于物料编码库（空前缀精确匹配）。"""
    if not material_code:
        return False
    exists = await session.scalar(
        select(MaterialCodeLibrary.id).where(MaterialCodeLibrary.material_code == material_code)
    )
    return exists is not None


async def search_material_codes(
    session: AsyncSession,
    *,
    keyword: str | None,
    name: str | None = None,
    model_spec: str | None = None,
    material_code: str | None = None,
    page: int,
    page_size: int,
) -> tuple[list[MaterialCodeLibraryRead], int]:
    query = select(MaterialCodeLibrary)
    condition = contains_any(
        (
            MaterialCodeLibrary.material_code,
            MaterialCodeLibrary.name,
            MaterialCodeLibrary.model_spec,
        ),
        keyword,
    )
    if condition is not None:
        query = query.where(condition)
    field_filters = (
        (MaterialCodeLibrary.name, name),
        (MaterialCodeLibrary.model_spec, model_spec),
        (MaterialCodeLibrary.material_code, material_code),
    )
    for column, value in field_filters:
        field_condition = contains_any((column,), value)
        if field_condition is not None:
            query = query.where(field_condition)
    total = int((await session.scalar(select(func.count()).select_from(query.subquery()))) or 0)
    result = await session.scalars(
        query.order_by(MaterialCodeLibrary.material_code)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    items = [
        MaterialCodeLibraryRead(
            id=item.id,
            material_code=item.material_code,
            name=item.name,
            model_spec=item.model_spec,
            unit_name=item.unit_name,
        )
        for item in result.all()
    ]
    return items, total
