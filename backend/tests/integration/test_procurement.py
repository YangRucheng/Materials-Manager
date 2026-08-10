from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from urllib.parse import unquote

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from PIL import Image
from sqlalchemy import select

from app.core.config import settings
from app.core.database import SessionLocal
from app.models import PurchaseMaterial, PurchaseRequest, PurchaseRequestLine
from tests.conftest import auth_headers, create_stock


async def create_purchase_plan(
    client: AsyncClient,
    headers: dict[str, str],
    name: str,
    *,
    code: str | None = None,
    category: str = "备品备件",
    stock_material_id: int | None = None,
    planned_qty: str = "5",
    model_spec: str = "M60-2P 5A",
    actual_demand_person: str = "车间员工张三",
    purchase_responsible: str = "李工",
    subitem_no: str | None = "01-01",
    plan_date: str = "2026-07-01",
    image_ids: list[str] | None = None,
) -> dict[str, object]:
    response = await client.post(
        "/api/v1/purchase-materials",
        headers=headers,
        json={
            "plan_date": plan_date,
            "material_code": code,
            "category": category,
            "name": name,
            "model_spec": model_spec,
            "unit_name": "个",
            "actual_demand_person": actual_demand_person,
            "purchase_responsible": purchase_responsible,
            "planned_qty": planned_qty,
            "usage": "控制柜检修",
            "subitem_no": subitem_no,
            "remark": "新计划",
            "stock_material_id": stock_material_id,
            "image_ids": image_ids or [],
        },
    )
    assert response.status_code == 201, response.text
    result = response.json()
    assert result["planned_qty"] == planned_qty
    assert result["purchase_responsible"] == purchase_responsible
    assert result["category"] == category
    assert result["plan_date"] == plan_date
    assert result["status"] == "正常"
    assert "code_state" not in result
    return result


@pytest.mark.asyncio
async def test_purchase_plan_status_filter_and_batch_archive(client: AsyncClient) -> None:
    purchase_headers = await auth_headers(client, "purchase")
    admin_headers = await auth_headers(client, "admin")
    normal = await create_purchase_plan(client, purchase_headers, "正常申购计划")
    deferred = await create_purchase_plan(client, purchase_headers, "暂不申购计划")
    archived = await create_purchase_plan(client, purchase_headers, "待归档申购计划")

    deferred_response = await client.patch(
        "/api/v1/purchase-materials/batch",
        headers=purchase_headers,
        json={
            "materials": [{"id": deferred["id"], "version": deferred["version"]}],
            "status": "暂不申购",
        },
    )
    assert deferred_response.status_code == 200, deferred_response.text
    assert deferred_response.json()[0]["status"] == "暂不申购"

    archive_response = await client.patch(
        "/api/v1/purchase-materials/batch",
        headers=purchase_headers,
        json={
            "materials": [{"id": archived["id"], "version": archived["version"]}],
            "status": "已归档",
        },
    )
    assert archive_response.status_code == 200, archive_response.text
    assert archive_response.json()[0]["status"] == "已归档"

    normal_list = await client.get(
        "/api/v1/purchase-materials",
        headers=purchase_headers,
        params={"status": "正常"},
    )
    deferred_list = await client.get(
        "/api/v1/purchase-materials",
        headers=purchase_headers,
        params={"status": "暂不申购"},
    )
    multi_status_list = await client.get(
        "/api/v1/purchase-materials",
        headers=purchase_headers,
        params=[("status", "正常"), ("status", "暂不申购")],
    )
    archived_list = await client.get(
        "/api/v1/purchase-materials",
        headers=purchase_headers,
        params={"status": "已归档"},
    )
    purchase_list = await client.get("/api/v1/purchase-materials", headers=purchase_headers)
    admin_archived_list = await client.get(
        "/api/v1/purchase-materials",
        headers=admin_headers,
        params={"status": "已归档"},
    )
    admin_all_list = await client.get("/api/v1/purchase-materials", headers=admin_headers)

    assert {item["id"] for item in normal_list.json()["items"]} == {normal["id"]}
    assert {item["id"] for item in deferred_list.json()["items"]} == {deferred["id"]}
    assert {item["id"] for item in multi_status_list.json()["items"]} == {
        normal["id"],
        deferred["id"],
    }
    assert archived_list.status_code == 403
    assert archived_list.json()["code"] == "ARCHIVED_PURCHASE_PLAN_FORBIDDEN"
    assert {item["id"] for item in purchase_list.json()["items"]} == {normal["id"]}
    assert {item["id"] for item in admin_archived_list.json()["items"]} == {archived["id"]}
    assert {item["id"] for item in admin_all_list.json()["items"]} == {
        normal["id"],
        deferred["id"],
        archived["id"],
    }
    forbidden_detail = await client.get(
        f"/api/v1/purchase-materials/{archived['id']}", headers=purchase_headers
    )
    assert forbidden_detail.status_code == 403
    admin_detail = await client.get(
        f"/api/v1/purchase-materials/{archived['id']}", headers=admin_headers
    )
    assert admin_detail.status_code == 200

    invalid = await client.get(
        "/api/v1/purchase-materials",
        headers=purchase_headers,
        params={"status": "已取消"},
    )
    assert invalid.status_code == 422


@pytest.mark.asyncio
async def test_empty_purchase_people_use_backslash_placeholder(client: AsyncClient) -> None:
    headers = await auth_headers(client, "purchase")
    response = await client.post(
        "/api/v1/purchase-materials",
        headers=headers,
        json={
            "plan_date": "2026-07-21",
            "name": "空负责人测试",
            "model_spec": "TEST-EMPTY",
            "unit_name": "个",
            "planned_qty": "1",
            "usage": "空值占位符测试",
            "image_ids": [],
        },
    )

    assert response.status_code == 201, response.text
    result = response.json()
    assert result["actual_demand_person"] == "\\"
    assert result["purchase_responsible"] == "\\"

    empty_people = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"empty_actual_demand_person": True, "moved": False},
    )
    assert empty_people.status_code == 200, empty_people.text
    assert [item["id"] for item in empty_people.json()["items"]] == [result["id"]]

    filter_options = await client.get(
        "/api/v1/purchase-materials/filter-options",
        headers=headers,
        params={"moved": False},
    )
    assert filter_options.status_code == 200, filter_options.text
    assert "\\" not in filter_options.json()["actual_demand_persons"]


@pytest.mark.asyncio
async def test_purchase_plan_subitem_filters_support_all_exact_and_empty(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(client, headers, "一号子项计划", subitem_no="01-01")
    second = await create_purchase_plan(client, headers, "二号子项计划", subitem_no="02-02")
    empty = await create_purchase_plan(client, headers, "空子项计划", subitem_no=None)

    options = await client.get(
        "/api/v1/purchase-materials/filter-options",
        headers=headers,
        params={"moved": False},
    )
    assert options.status_code == 200, options.text
    assert options.json()["subitem_nos"] == ["01-01", "02-02"]

    exact = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"moved": False, "subitem_no": "02-02"},
    )
    assert [item["id"] for item in exact.json()["items"]] == [second["id"]]

    empty_result = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"moved": False, "empty_subitem_no": True},
    )
    assert [item["id"] for item in empty_result.json()["items"]] == [empty["id"]]

    all_result = await client.get(
        "/api/v1/purchase-materials", headers=headers, params={"moved": False}
    )
    assert {item["id"] for item in all_result.json()["items"]} == {
        first["id"],
        second["id"],
        empty["id"],
    }


@pytest.mark.asyncio
async def test_purchase_lists_support_field_like_and_person_filters(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(
        client,
        headers,
        "智能断路器",
        code="DQ-FILTER-001",
        model_spec="DZ47-2P 5A",
        actual_demand_person="张三",
        purchase_responsible="李工",
    )
    second = await create_purchase_plan(
        client,
        headers,
        "交流接触器",
        code="DQ-FILTER-002",
        model_spec="CJX2-18 220V",
        actual_demand_person="李四",
        purchase_responsible="王工",
    )

    plan_search = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"search_field": "name", "search_value": "断路", "moved": False},
    )
    assert plan_search.status_code == 200, plan_search.text
    assert [item["id"] for item in plan_search.json()["items"]] == [first["id"]]

    person_search = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"purchase_responsible": "王", "moved": False},
    )
    assert person_search.status_code == 200, person_search.text
    assert [item["id"] for item in person_search.json()["items"]] == [second["id"]]

    combined_plan_search = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={
            "name": "断路",
            "model_spec": "2P",
            "purchase_responsible": "李",
            "moved": False,
        },
    )
    assert combined_plan_search.status_code == 200, combined_plan_search.text
    assert [item["id"] for item in combined_plan_search.json()["items"]] == [first["id"]]

    mismatched_plan_search = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"name": "断路", "purchase_responsible": "王", "moved": False},
    )
    assert mismatched_plan_search.status_code == 200, mismatched_plan_search.text
    assert mismatched_plan_search.json()["items"] == []

    plan_options = await client.get(
        "/api/v1/purchase-materials/filter-options",
        headers=headers,
        params={"moved": False},
    )
    assert plan_options.status_code == 200, plan_options.text
    assert set(plan_options.json()["actual_demand_persons"]) == {"张三", "李四"}
    assert set(plan_options.json()["purchase_responsibles"]) == {"李工", "王工"}

    first_record = await move_to_record(
        client, headers, int(first["id"]), trace_no="TRACE-LIKE-001"
    )
    second_record = await move_to_record(
        client,
        headers,
        int(second["id"]),
        trace_no="TRACE-OTHER-002",
        salesperson="钱经理",
    )
    async with SessionLocal() as session:
        line = await session.get(PurchaseRequestLine, int(first_record["line_id"]))
        assert line is not None
        line.status = ""
        await session.commit()

    record_search = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={
            "search_field": "trace_no",
            "search_value": "LIKE-00",
            "actual_demand_person": "张",
            "page_size": 10,
        },
    )
    assert record_search.status_code == 200, record_search.text
    assert record_search.json()["page_size"] == 10
    assert [item["purchase_material_id"] for item in record_search.json()["items"]] == [first["id"]]

    combined_record_search = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={
            "name": "接触",
            "model_spec": "220V",
            "purchase_responsible": "王",
            "salesperson": "钱",
        },
    )
    assert combined_record_search.status_code == 200, combined_record_search.text
    assert [item["purchase_material_id"] for item in combined_record_search.json()["items"]] == [
        second["id"]
    ]

    salesperson_search = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"salesperson": "钱"},
    )
    assert salesperson_search.status_code == 200, salesperson_search.text
    assert [item["line_id"] for item in salesperson_search.json()["items"]] == [
        second_record["line_id"]
    ]

    number_search = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={
            "purchase_order_no": "SG-2026",
            "trace_no": "OTHER-002",
        },
    )
    assert number_search.status_code == 200, number_search.text
    assert [item["purchase_material_id"] for item in number_search.json()["items"]] == [
        second["id"]
    ]

    mismatched_number_search = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={
            "purchase_order_no": "SG-2026",
            "trace_no": "LIKE-001",
            "name": "接触器",
        },
    )
    assert mismatched_number_search.status_code == 200, mismatched_number_search.text
    assert mismatched_number_search.json()["items"] == []

    record_options = await client.get("/api/v1/purchase-records/filter-options", headers=headers)
    assert record_options.status_code == 200, record_options.text
    assert set(record_options.json()["actual_demand_persons"]) == {"张三", "李四"}
    assert set(record_options.json()["salespersons"]) == {"赵经理", "钱经理"}
    assert record_options.json()["statuses"] == ["已申购"]

    selected_status = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"status": "已申购"},
    )
    assert selected_status.status_code == 200, selected_status.text
    assert [item["line_id"] for item in selected_status.json()["items"]] == [
        second_record["line_id"]
    ]

    empty_status = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"empty_status": True},
    )
    assert empty_status.status_code == 200, empty_status.text
    assert [item["line_id"] for item in empty_status.json()["items"]] == [first_record["line_id"]]


async def move_to_record(
    client: AsyncClient,
    headers: dict[str, str],
    plan_id: int,
    trace_no: str | None = "ZS-2026-001",
    salesperson: str = "赵经理",
    purchase_order_no: str | None = "SG-2026-001",
) -> dict[str, object]:
    response = await client.post(
        f"/api/v1/purchase-materials/{plan_id}/move-to-record",
        headers=headers,
        json={
            "purchase_order_no": purchase_order_no,
            "trace_no": trace_no,
            "contract_no": "HT-2026-001",
            "vessel_no": "VESSEL-01",
            "consolidation_date": "2026-07-19",
            "consolidation_port": "上海港",
            "sailing_date": "2026-07-20",
            "purchase_date": "2026-07-18",
            "salesperson": salesperson,
            "status": "已申购",
            "record_remark": "供应商信息待补充",
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


@pytest.mark.asyncio
async def test_purchase_records_default_to_purchase_and_trace_number_order(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    order_numbers = (
        ("ORDER-B", "TRACE-A"),
        ("ORDER-A", "TRACE-C"),
        ("ORDER-A", None),
        ("ORDER-A", "TRACE-A"),
        (None, "TRACE-Z"),
    )
    for index, (purchase_order_no, trace_no) in enumerate(order_numbers, start=1):
        plan = await create_purchase_plan(
            client,
            headers,
            f"追溯号排序物料-{index}",
            code=f"TRACE-SORT-{index}",
        )
        await move_to_record(
            client,
            headers,
            int(plan["id"]),
            trace_no=trace_no,
            purchase_order_no=purchase_order_no,
        )

    response = await client.get("/api/v1/purchase-records", headers=headers)

    assert response.status_code == 200, response.text
    assert [
        (item["purchase_order_no"], item["trace_no"])
        for item in response.json()["items"]
    ] == [
        ("ORDER-B", "TRACE-A"),
        ("ORDER-A", "TRACE-C"),
        ("ORDER-A", "TRACE-A"),
        ("ORDER-A", None),
        (None, "TRACE-Z"),
    ]

    first_page = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"page": 1, "page_size": 2},
    )
    assert first_page.status_code == 200, first_page.text
    assert [
        (item["purchase_order_no"], item["trace_no"])
        for item in first_page.json()["items"]
    ] == [
        ("ORDER-B", "TRACE-A"),
        ("ORDER-A", "TRACE-C"),
    ]


@pytest.mark.asyncio
async def test_purchase_plan_list_supports_arbitrary_column_sorting(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    zebra = await create_purchase_plan(
        client, headers, "Zebra", code="SORT-Z-01", actual_demand_person="甲"
    )
    alpha = await create_purchase_plan(
        client, headers, "Alpha", code="SORT-A-02", actual_demand_person="乙"
    )
    bravo = await create_purchase_plan(
        client, headers, "Bravo", code="SORT-B-03", actual_demand_person="丙"
    )

    asc_response = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"sort_by": "name", "sort_order": "asc", "moved": False},
    )
    assert asc_response.status_code == 200, asc_response.text
    assert [item["name"] for item in asc_response.json()["items"]] == [
        "Alpha",
        "Bravo",
        "Zebra",
    ]

    desc_response = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"sort_by": "name", "sort_order": "desc", "moved": False},
    )
    assert desc_response.status_code == 200, desc_response.text
    assert [item["name"] for item in desc_response.json()["items"]] == [
        "Zebra",
        "Bravo",
        "Alpha",
    ]

    default_response = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"moved": False},
    )
    assert default_response.status_code == 200, default_response.text
    # 未传 sort_by 时保持默认 id 倒序
    assert [item["id"] for item in default_response.json()["items"]] == [
        bravo["id"],
        alpha["id"],
        zebra["id"],
    ]


@pytest.mark.asyncio
async def test_purchase_record_list_supports_arbitrary_column_sorting(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(client, headers, "排序物料甲", code="REC-SORT-01")
    second = await create_purchase_plan(client, headers, "排序物料乙", code="REC-SORT-02")
    third = await create_purchase_plan(client, headers, "排序物料丙", code="REC-SORT-03")
    await move_to_record(
        client,
        headers,
        int(first["id"]),
        purchase_order_no="SG-SORT-001",
        trace_no="ZS-SORT-001",
        salesperson="王业务",
    )
    await move_to_record(
        client,
        headers,
        int(second["id"]),
        purchase_order_no="SG-SORT-003",
        trace_no="ZS-SORT-003",
        salesperson="张业务",
    )
    await move_to_record(
        client,
        headers,
        int(third["id"]),
        purchase_order_no="SG-SORT-002",
        trace_no="ZS-SORT-002",
        salesperson="李业务",
    )

    desc_response = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"sort_by": "purchase_order_no", "sort_order": "desc"},
    )
    assert desc_response.status_code == 200, desc_response.text
    assert [item["purchase_order_no"] for item in desc_response.json()["items"]] == [
        "SG-SORT-003",
        "SG-SORT-002",
        "SG-SORT-001",
    ]

    name_asc = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"sort_by": "purchase_order_no", "sort_order": "asc"},
    )
    assert name_asc.status_code == 200, name_asc.text
    assert [item["purchase_order_no"] for item in name_asc.json()["items"]] == [
        "SG-SORT-001",
        "SG-SORT-002",
        "SG-SORT-003",
    ]


@pytest.mark.asyncio
async def test_purchase_list_sort_by_whitelist_rejects_unknown_column(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    plan_response = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"sort_by": "not_a_column", "moved": False},
    )
    assert plan_response.status_code == 422, plan_response.text

    record_response = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"sort_by": "not_a_column"},
    )
    assert record_response.status_code == 422, record_response.text


@pytest.mark.asyncio
async def test_purchase_search_supports_or_delimiters_and_keeps_filters_anded(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(
        client,
        headers,
        "智能断路器",
        code="DQ-OR-001",
        actual_demand_person="张三",
        purchase_responsible="李工",
    )
    second = await create_purchase_plan(
        client,
        headers,
        "交流接触器",
        code="DQ-OR-002",
        actual_demand_person="李四",
        purchase_responsible="王工",
    )
    literal = await create_purchase_plan(
        client,
        headers,
        "100%负载开关",
        code="DQ-OR-003",
        actual_demand_person="赵六",
        purchase_responsible="周工",
    )

    plan_field_search = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"search_field": "name", "search_value": "断路||接触器|", "moved": False},
    )
    assert plan_field_search.status_code == 200, plan_field_search.text
    assert {item["id"] for item in plan_field_search.json()["items"]} == {
        first["id"],
        second["id"],
    }

    plan_fullwidth_search = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"name": "断路｜接触器", "moved": False},
    )
    assert plan_fullwidth_search.status_code == 200, plan_fullwidth_search.text
    assert {item["id"] for item in plan_fullwidth_search.json()["items"]} == {
        first["id"],
        second["id"],
    }

    plan_global_search = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"keyword": "DQ-OR-001|王工", "moved": False},
    )
    assert plan_global_search.status_code == 200, plan_global_search.text
    assert {item["id"] for item in plan_global_search.json()["items"]} == {
        first["id"],
        second["id"],
    }

    plan_anded_search = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"name": "断路|接触器", "purchase_responsible": "李", "moved": False},
    )
    assert plan_anded_search.status_code == 200, plan_anded_search.text
    assert [item["id"] for item in plan_anded_search.json()["items"]] == [first["id"]]

    literal_search = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"name": "%", "moved": False},
    )
    assert literal_search.status_code == 200, literal_search.text
    assert [item["id"] for item in literal_search.json()["items"]] == [literal["id"]]

    first_record = await move_to_record(
        client, headers, int(first["id"]), trace_no="TRACE-OR-001", salesperson="赵经理"
    )
    second_record = await move_to_record(
        client, headers, int(second["id"]), trace_no="TRACE-OR-002", salesperson="钱经理"
    )

    record_field_search = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"search_field": "trace_no", "search_value": "OR-001｜OR-002"},
    )
    assert record_field_search.status_code == 200, record_field_search.text
    assert {item["line_id"] for item in record_field_search.json()["items"]} == {
        first_record["line_id"],
        second_record["line_id"],
    }

    record_trace_search = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"trace_no": "OR-001||OR-002|"},
    )
    assert record_trace_search.status_code == 200, record_trace_search.text
    assert {item["line_id"] for item in record_trace_search.json()["items"]} == {
        first_record["line_id"],
        second_record["line_id"],
    }

    record_global_search = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"keyword": "DQ-OR-001||钱经理|"},
    )
    assert record_global_search.status_code == 200, record_global_search.text
    assert {item["line_id"] for item in record_global_search.json()["items"]} == {
        first_record["line_id"],
        second_record["line_id"],
    }

    record_anded_search = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"name": "断路|接触器", "salesperson": "钱"},
    )
    assert record_anded_search.status_code == 200, record_anded_search.text
    assert [item["line_id"] for item in record_anded_search.json()["items"]] == [
        second_record["line_id"]
    ]


@pytest.mark.asyncio
async def test_plan_number_uses_date_sequence_and_record_keeps_plan_date(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(client, headers, "同日计划一", code="DQ-PLAN-1")
    second = await create_purchase_plan(client, headers, "同日计划二", code="DQ-PLAN-2")
    other_day = await create_purchase_plan(
        client,
        headers,
        "次日计划",
        code="DQ-PLAN-3",
        plan_date="2026-07-02",
    )

    assert first["plan_no"] == "PLAN-20260701-001"
    assert second["plan_no"] == "PLAN-20260701-002"
    assert other_day["plan_no"] == "PLAN-20260702-001"

    record = await move_to_record(client, headers, int(first["id"]))
    assert record["plan_no"] == first["plan_no"]
    assert record["plan_date"] == "2026-07-01"
    assert record["category"] == "备品备件"
    assert record["contract_no"] == "HT-2026-001"
    assert record["vessel_no"] == "VESSEL-01"
    assert record["consolidation_date"] == "2026-07-19"
    assert record["consolidation_port"] == "上海港"
    assert record["sailing_date"] == "2026-07-20"


@pytest.mark.asyncio
async def test_record_keeps_purchase_plan_attachments(client: AsyncClient) -> None:
    headers = await auth_headers(client, "purchase")
    source = BytesIO()
    Image.new("RGB", (24, 16), "blue").save(source, format="PNG")
    uploaded = await client.post(
        "/api/v1/files/images",
        headers=headers,
        files={"file": ("motor.png", source.getvalue(), "image/png")},
    )
    assert uploaded.status_code == 201, uploaded.text
    file_id = uploaded.json()["id"]
    plan = await create_purchase_plan(
        client,
        headers,
        "带附件申购计划",
        code="DQ-IMAGE-001",
        image_ids=[file_id],
    )

    record = await move_to_record(client, headers, int(plan["id"]))

    assert [image["id"] for image in record["images"]] == [file_id]
    detail = await client.get(f"/api/v1/purchase-records/{record['line_id']}", headers=headers)
    assert detail.status_code == 200, detail.text
    assert [image["id"] for image in detail.json()["images"]] == [file_id]

    from app.services import purchase_plan_cleanup_service

    deleted = await purchase_plan_cleanup_service.cleanup_moved_plans_once()
    assert deleted == 1
    detail = await client.get(f"/api/v1/purchase-records/{record['line_id']}", headers=headers)
    assert detail.status_code == 200, detail.text
    assert [image["id"] for image in detail.json()["images"]] == [file_id]


@pytest.mark.asyncio
async def test_uncoded_plan_must_be_coded_before_moving_to_record(client: AsyncClient) -> None:
    headers = await auth_headers(client, "purchase")
    plan = await create_purchase_plan(client, headers, "未编码物资")

    uncoded = await client.get("/api/v1/purchase-materials?coded=false", headers=headers)
    assert uncoded.status_code == 200
    assert [item["id"] for item in uncoded.json()["items"]] == [plan["id"]]

    rejected = await client.post(
        f"/api/v1/purchase-materials/{plan['id']}/move-to-record",
        headers=headers,
        json={
            "purchase_order_no": "SG-UNCODED",
            "trace_no": "ZS-UNCODED",
            "purchase_date": "2026-07-18",
        },
    )
    assert rejected.status_code == 409
    assert rejected.json()["code"] == "MATERIAL_CODE_REQUIRED"

    coded = await client.patch(
        f"/api/v1/purchase-materials/{plan['id']}",
        headers=headers,
        json={
            "version": plan["version"],
            "material_code": "DQ-000500",
            "name": plan["name"],
            "model_spec": plan["model_spec"],
            "unit_name": plan["unit_name"],
            "actual_demand_person": plan["actual_demand_person"],
            "purchase_responsible": plan["purchase_responsible"],
            "planned_qty": plan["planned_qty"],
            "usage": plan["usage"],
            "subitem_no": plan["subitem_no"],
            "remark": plan["remark"],
            "stock_material_id": None,
            "image_ids": [],
        },
    )
    assert coded.status_code == 200, coded.text
    record = await move_to_record(client, headers, int(plan["id"]))
    assert record["material_code"] == "DQ-000500"
    assert record["purchase_qty"] == "5"
    assert record["status"] == "已申购"


@pytest.mark.asyncio
async def test_plan_allows_optional_subitem_and_manual_stock_link(client: AsyncClient) -> None:
    purchase = await auth_headers(client, "purchase")
    warehouse = await auth_headers(client, "warehouse")
    stock_id = await create_stock(client, warehouse, "手动关联物资")

    plan = await create_purchase_plan(
        client,
        purchase,
        "手动关联物资",
        code="DQ-LINK-001",
        stock_material_id=stock_id,
        subitem_no=None,
    )

    assert plan["subitem_no"] is None
    assert plan["stock_material_id"] == stock_id
    assert plan["stock_material_name"] == "手动关联物资"
    record = await move_to_record(client, purchase, int(plan["id"]), "无子项申购")
    assert record["subitem_no"] is None


@pytest.mark.asyncio
async def test_plan_can_be_deleted_until_moved_to_record(client: AsyncClient) -> None:
    headers = await auth_headers(client, "purchase")
    deletable = await create_purchase_plan(client, headers, "待删除计划")
    deleted = await client.delete(
        f"/api/v1/purchase-materials/{deletable['id']}",
        headers={**headers, "If-Match": str(deletable["version"])},
    )
    assert deleted.status_code == 204, deleted.text
    missing = await client.get(f"/api/v1/purchase-materials/{deletable['id']}", headers=headers)
    assert missing.status_code == 400

    moved = await create_purchase_plan(client, headers, "已转入计划", code="DQ-MOVED")
    await move_to_record(client, headers, int(moved["id"]))
    rejected = await client.delete(
        f"/api/v1/purchase-materials/{moved['id']}",
        headers={**headers, "If-Match": str(moved["version"])},
    )
    assert rejected.status_code == 409
    assert rejected.json()["code"] == "PURCHASE_PLAN_IN_USE"


@pytest.mark.asyncio
async def test_purchase_record_can_restore_to_purchase_plan(client: AsyncClient) -> None:
    headers = await auth_headers(client, "purchase")
    plan = await create_purchase_plan(
        client,
        headers,
        "被打回的申购物资",
        code="DQ-RESTORE-1",
        category="工具",
    )
    record = await move_to_record(client, headers, int(plan["id"]))

    restored = await client.post(
        f"/api/v1/purchase-records/{record['line_id']}/restore-to-plan",
        headers={**headers, "If-Match": str(record["version"])},
    )
    assert restored.status_code == 200, restored.text
    restored_plan = restored.json()
    assert restored_plan["id"] == plan["id"]
    assert restored_plan["category"] == "工具"
    assert restored_plan["status"] == "正常"
    assert restored_plan["moved_to_record"] is False
    assert "contract_no" not in restored_plan
    assert "purchase_date" not in restored_plan

    missing_record = await client.get(
        f"/api/v1/purchase-records/{record['line_id']}", headers=headers
    )
    assert missing_record.status_code == 400

    async with SessionLocal() as session:
        request = await session.get(PurchaseRequest, int(record["purchase_request_id"]))
        assert request is None


@pytest.mark.asyncio
async def test_restore_rebuilds_plan_after_cleanup(client: AsyncClient) -> None:
    """计划被定时清理后，恢复为申购计划从快照重建新计划但保留原计划号。"""
    from app.services import purchase_plan_cleanup_service

    headers = await auth_headers(client, "purchase")
    plan = await create_purchase_plan(
        client, headers, "已清理后恢复", code="DQ-RESTORE-CLEANED", category="工具"
    )
    record = await move_to_record(client, headers, int(plan["id"]))

    deleted = await purchase_plan_cleanup_service.cleanup_moved_plans_once()
    assert deleted == 1
    async with SessionLocal() as session:
        # 清理后记录行已解绑外键（指向已删计划）
        line = await session.get(PurchaseRequestLine, int(record["line_id"]))
        assert line is not None
        assert line.purchase_material_id is None

    restored = await client.post(
        f"/api/v1/purchase-records/{record['line_id']}/restore-to-plan",
        headers={**headers, "If-Match": str(record["version"])},
    )
    assert restored.status_code == 200, restored.text
    restored_plan = restored.json()
    assert restored_plan["plan_no"] == plan["plan_no"]  # 保留原计划号
    assert restored_plan["category"] == "工具"
    assert restored_plan["status"] == "正常"
    assert restored_plan["moved_to_record"] is False
    async with SessionLocal() as session:
        # 恢复后重建出独立存在的计划，保留原计划号
        rebuilt = await session.scalar(
            select(PurchaseMaterial).where(PurchaseMaterial.plan_no == plan["plan_no"])
        )
        assert rebuilt is not None
        assert rebuilt.id == int(restored_plan["id"])


@pytest.mark.asyncio
async def test_multiple_plans_can_move_to_one_purchase_record_batch(client: AsyncClient) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(client, headers, "批量计划一", code="DQ-BATCH-1")
    second = await create_purchase_plan(
        client,
        headers,
        "批量计划二",
        code="DQ-BATCH-2",
        purchase_responsible="另一负责人",
    )
    response = await client.post(
        "/api/v1/purchase-materials/batch-move-to-record",
        headers=headers,
        json={
            "material_ids": [first["id"], second["id"]],
            "purchase_order_no": "SG-BATCH-001",
            "trace_no": "ZS-BATCH-001",
            "purchase_date": "2026-07-18",
            "salesperson": "批量业务员",
            "status": "采购处理中",
            "record_remark": "批量转入",
        },
    )
    assert response.status_code == 200, response.text
    records = response.json()
    assert len(records) == 2
    assert {record["purchase_material_id"] for record in records} == {
        first["id"],
        second["id"],
    }
    assert len({record["purchase_request_id"] for record in records}) == 1
    assert {record["trace_no"] for record in records} == {"ZS-BATCH-001"}
    assert {record["purchase_order_no"] for record in records} == {"SG-BATCH-001"}
    assert {record["status"] for record in records} == {"采购处理中"}

    restored = await client.post(
        f"/api/v1/purchase-records/{records[0]['line_id']}/restore-to-plan",
        headers={**headers, "If-Match": str(records[0]["version"])},
    )
    assert restored.status_code == 200, restored.text
    assert restored.json()["moved_to_record"] is False

    remaining = await client.get(
        f"/api/v1/purchase-records/{records[1]['line_id']}", headers=headers
    )
    assert remaining.status_code == 200, remaining.text
    assert remaining.json()["purchase_request_id"] == records[1]["purchase_request_id"]
    assert remaining.json()["version"] == records[1]["version"] + 1


@pytest.mark.asyncio
async def test_batch_update_purchase_plans(client: AsyncClient) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(client, headers, "批量修改计划一", subitem_no="01-01")
    second = await create_purchase_plan(
        client,
        headers,
        "批量修改计划二",
        subitem_no="02-02",
        actual_demand_person="原需求人",
    )

    response = await client.patch(
        "/api/v1/purchase-materials/batch",
        headers=headers,
        json={
            "materials": [
                {"id": first["id"], "version": first["version"]},
                {"id": second["id"], "version": second["version"]},
            ],
            "plan_date": "2026-07-15",
            "category": "工具",
            "urgency": "非常紧急",
            "demand_department": "HXNI 设备管理部",
            "actual_demand_person": "统一需求人",
            "purchase_responsible": "统一申购负责人",
            "subitem_no": None,
            "usage": "统一批量修改用途",
        },
    )

    assert response.status_code == 200, response.text
    payload = response.json()
    assert [item["id"] for item in payload] == [first["id"], second["id"]]
    assert {item["plan_date"] for item in payload} == {"2026-07-15"}
    assert {item["category"] for item in payload} == {"工具"}
    assert {item["urgency"] for item in payload} == {"非常紧急"}
    assert {item["demand_department"] for item in payload} == {"HXNI 设备管理部"}
    assert {item["actual_demand_person"] for item in payload} == {"统一需求人"}
    assert {item["purchase_responsible"] for item in payload} == {"统一申购负责人"}
    assert {item["subitem_no"] for item in payload} == {None}
    assert {item["usage"] for item in payload} == {"统一批量修改用途"}
    assert len({item["plan_no"] for item in payload}) == 2
    assert all(item["plan_no"].startswith("PLAN-20260715-") for item in payload)
    assert {item["version"] for item in payload} == {2}

    listed = await client.get(
        "/api/v1/purchase-materials",
        headers=headers,
        params={"moved": False, "actual_demand_person": "统一需求人"},
    )
    assert listed.status_code == 200, listed.text
    assert {item["id"] for item in listed.json()["items"]} == {first["id"], second["id"]}


@pytest.mark.asyncio
async def test_batch_update_purchase_records_is_atomic(client: AsyncClient) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(client, headers, "批量修改记录一", code="DQ-REC-BATCH-1")
    second = await create_purchase_plan(client, headers, "批量修改记录二", code="DQ-REC-BATCH-2")
    moved = await client.post(
        "/api/v1/purchase-materials/batch-move-to-record",
        headers=headers,
        json={
            "material_ids": [first["id"], second["id"]],
            "purchase_order_no": "SG-REC-BEFORE",
            "trace_no": "ZS-REC-BEFORE",
            "purchase_date": "2026-07-18",
            "status": "已申购",
        },
    )
    assert moved.status_code == 200, moved.text
    records = moved.json()

    changed = await client.patch(
        "/api/v1/purchase-records/batch",
        headers=headers,
        json={
            "records": [
                {"line_id": record["line_id"], "version": record["version"]}
                for record in records
            ],
            "plan_date": "2026-07-19",
            "purchase_order_no": "SG-REC-AFTER",
            "trace_no": "ZS-REC-AFTER",
            "contract_no": "HT-REC-BATCH",
            "purchase_date": "2026-07-20",
            "actual_demand_person": "统一实际需求人",
            "purchase_responsible": "统一申购负责人",
            "salesperson": "统一业务员",
            "status": "批量处理中",
            "record_remark": "批量修改记录备注",
        },
    )
    assert changed.status_code == 200, changed.text
    payload = changed.json()
    assert [item["line_id"] for item in payload] == [item["line_id"] for item in records]
    assert {item["plan_date"] for item in payload} == {"2026-07-19"}
    assert len({item["plan_no"] for item in payload}) == 2
    assert all(item["plan_no"].startswith("PLAN-20260719-") for item in payload)
    assert {item["purchase_order_no"] for item in payload} == {"SG-REC-AFTER"}
    assert {item["trace_no"] for item in payload} == {"ZS-REC-AFTER"}
    assert {item["contract_no"] for item in payload} == {"HT-REC-BATCH"}
    assert {item["purchase_date"] for item in payload} == {"2026-07-20"}
    assert {item["actual_demand_person"] for item in payload} == {"统一实际需求人"}
    assert {item["purchase_responsible"] for item in payload} == {"统一申购负责人"}
    assert {item["salesperson"] for item in payload} == {"统一业务员"}
    assert {item["status"] for item in payload} == {"批量处理中"}
    assert {item["record_remark"] for item in payload} == {"批量修改记录备注"}
    assert {item["version"] for item in payload} == {records[0]["version"] + 1}

    conflict = await client.patch(
        "/api/v1/purchase-records/batch",
        headers=headers,
        json={
            "records": [
                {"line_id": payload[0]["line_id"], "version": payload[0]["version"]},
                {"line_id": payload[1]["line_id"], "version": records[1]["version"]},
            ],
            "status": "不应部分生效",
        },
    )
    assert conflict.status_code == 409, conflict.text

    unchanged = await client.get(
        f"/api/v1/purchase-records/{payload[0]['line_id']}", headers=headers
    )
    assert unchanged.status_code == 200, unchanged.text
    assert unchanged.json()["status"] == "批量处理中"


@pytest.mark.asyncio
async def test_grouped_purchase_request_supports_line_tracking_fields(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(client, headers, "行级追溯一", code="DQ-LINE-TRACK-1")
    second = await create_purchase_plan(client, headers, "行级追溯二", code="DQ-LINE-TRACK-2")
    moved = await client.post(
        "/api/v1/purchase-materials/batch-move-to-record",
        headers=headers,
        json={
            "material_ids": [first["id"], second["id"]],
            "purchase_order_no": "SG-LINE-TRACK",
            "trace_no": "HEADER-TRACE",
            "purchase_date": "2026-08-01",
            "salesperson": "主表业务员",
        },
    )
    assert moved.status_code == 200, moved.text
    records = moved.json()

    async with SessionLocal() as session:
        first_line = await session.get(PurchaseRequestLine, int(records[0]["line_id"]))
        second_line = await session.get(PurchaseRequestLine, int(records[1]["line_id"]))
        assert first_line is not None
        assert second_line is not None
        first_line.trace_no = "LINE-TRACE-1"
        first_line.salesperson = "行业务员一"
        second_line.trace_no = "LINE-TRACE-2"
        second_line.salesperson = "行业务员二"
        await session.commit()

    listed = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"purchase_order_no": "SG-LINE-TRACK", "page_size": 10},
    )
    assert listed.status_code == 200, listed.text
    values = {
        item["material_code"]: (item["trace_no"], item["salesperson"])
        for item in listed.json()["items"]
    }
    assert values == {
        "DQ-LINE-TRACK-1": ("LINE-TRACE-1", "行业务员一"),
        "DQ-LINE-TRACK-2": ("LINE-TRACE-2", "行业务员二"),
    }

    filtered = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"trace_no": "LINE-TRACE-2", "salesperson": "行业务员二"},
    )
    assert filtered.status_code == 200, filtered.text
    assert [item["material_code"] for item in filtered.json()["items"]] == [
        "DQ-LINE-TRACK-2"
    ]


@pytest.mark.asyncio
async def test_purchase_record_supports_full_edit_and_free_text_status(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    first = await create_purchase_plan(
        client, headers, "负责人一", code="DQ-1", purchase_responsible="外部负责人甲"
    )
    second = await create_purchase_plan(
        client, headers, "负责人二", code="DQ-2", purchase_responsible="外部负责人乙"
    )
    first_record = await move_to_record(client, headers, int(first["id"]), "正式申购-A")
    await move_to_record(client, headers, int(second["id"]), "正式申购-B")

    remaining_plans = await client.get("/api/v1/purchase-materials?moved=false", headers=headers)
    assert remaining_plans.json()["total"] == 0

    records = await client.get("/api/v1/purchase-records", headers=headers)
    assert records.status_code == 200
    assert records.json()["total"] == 2
    assert len(records.json()["items"]) == 2
    assert all("material_name" in item for item in records.json()["items"])

    changed = await client.patch(
        f"/api/v1/purchase-records/{first_record['line_id']}",
        headers=headers,
        json={
            "version": first_record["version"],
            "plan_date": "2026-07-20",
            "material_code": "DQ-1-REV",
            "material_name": "负责人一修订",
            "model_spec": "M60-2P 10A",
            "unit_name": "个",
            "actual_demand_person": "检修班王五",
            "purchase_responsible": "外部负责人丙",
            "purchase_qty": "8",
            "usage": "统计修订用途",
            "subitem_no": "02-02",
            "plan_remark": "计划备注修订",
            "stock_material_id": None,
            "image_ids": [],
            "purchase_order_no": "SG-A-修订",
            "trace_no": "ZS-A-修订",
            "purchase_date": "2026-07-19",
            "salesperson": "钱经理",
            "status": "供应商已确认，等待财务安排",
            "record_remark": "仅用于整理统计",
        },
    )
    assert changed.status_code == 200, changed.text
    assert changed.json()["trace_no"] == "ZS-A-修订"
    assert changed.json()["purchase_order_no"] == "SG-A-修订"
    assert changed.json()["purchase_date"] == "2026-07-19"
    assert changed.json()["salesperson"] == "钱经理"
    assert changed.json()["plan_date"] == "2026-07-20"
    assert changed.json()["material_code"] == "DQ-1-REV"
    assert changed.json()["material_name"] == "负责人一修订"
    assert changed.json()["model_spec"] == "M60-2P 10A"
    assert changed.json()["actual_demand_person"] == "检修班王五"
    assert changed.json()["purchase_responsible"] == "外部负责人丙"
    assert changed.json()["purchase_qty"] == "8"
    assert changed.json()["usage"] == "统计修订用途"
    assert changed.json()["subitem_no"] == "02-02"
    assert changed.json()["plan_remark"] == "计划备注修订"
    assert changed.json()["status"] == "供应商已确认，等待财务安排"
    assert changed.json()["record_remark"] == "仅用于整理统计"
    assert "received_qty" not in changed.json()
    assert "remaining_qty" not in changed.json()

    async with SessionLocal() as session:
        material = await session.get(PurchaseMaterial, int(first["id"]))
        assert material is not None
        # 编辑记录不再回写计划：计划的名称/型号/数量保持转入时原样
        assert material.name == "负责人一"
        assert material.model_spec == "M60-2P 5A"
        assert str(material.planned_qty) == "5.0"

    filtered = await client.get(
        "/api/v1/purchase-records",
        headers=headers,
        params={"status": "供应商已确认，等待财务安排"},
    )
    assert filtered.status_code == 200, filtered.text
    assert [item["line_id"] for item in filtered.json()["items"]] == [first_record["line_id"]]

    duplicate = await client.post(
        f"/api/v1/purchase-materials/{first['id']}/move-to-record",
        headers=headers,
        json={
            "purchase_order_no": "SG-REPEAT",
            "trace_no": "ZS-REPEAT",
            "purchase_date": "2026-07-18",
        },
    )
    assert duplicate.status_code == 409
    assert duplicate.json()["code"] == "PLAN_ALREADY_MOVED"


@pytest.mark.asyncio
async def test_inventory_inbound_does_not_change_purchase_record(client: AsyncClient) -> None:
    purchase = await auth_headers(client, "purchase")
    warehouse = await auth_headers(client, "warehouse")
    stock_id = await create_stock(client, warehouse, "智能电机保护器")
    plan = await create_purchase_plan(
        client,
        purchase,
        "智能电机保护器",
        code="DQ-000501",
        stock_material_id=stock_id,
        planned_qty="5",
    )
    record = await move_to_record(client, purchase, int(plan["id"]))

    inbound = await client.post(
        "/api/v1/inventory/inbounds",
        headers=warehouse,
        json={
            "client_request_id": "independent-inbound",
            "occurred_at": "2026-07-17T10:00:00+08:00",
            "source_type": "MANUAL",
            "business_reason": "普通入库，与申购记录无关",
            "lines": [{"stock_material_id": stock_id, "quantity": "5"}],
        },
    )
    assert inbound.status_code == 201, inbound.text
    unchanged = await client.get(f"/api/v1/purchase-records/{record['line_id']}", headers=purchase)
    assert unchanged.status_code == 200, unchanged.text
    assert unchanged.json()["purchase_qty"] == "5"
    assert unchanged.json()["status"] == "已申购"
    assert "received_qty" not in unchanged.json()
    assert "remaining_qty" not in unchanged.json()


@pytest.mark.asyncio
async def test_purchase_plans_can_repeat_code_and_stock_material(client: AsyncClient) -> None:
    purchase = await auth_headers(client, "purchase")
    warehouse = await auth_headers(client, "warehouse")
    stock_id = await create_stock(client, warehouse, "重复补库物资")
    first = await create_purchase_plan(
        client, purchase, "重复补库物资", code="DQ-REPEAT", stock_material_id=stock_id
    )
    second = await create_purchase_plan(
        client, purchase, "重复补库物资", code="DQ-REPEAT", stock_material_id=stock_id
    )
    assert first["id"] != second["id"]


@pytest.mark.asyncio
async def test_purchase_tracking_numbers_are_optional_and_order_number_defaults(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    blank = await create_purchase_plan(client, headers, "空单号计划", code="DQ-BLANK")
    response = await client.post(
        f"/api/v1/purchase-materials/{blank['id']}/move-to-record",
        headers=headers,
        json={
            "purchase_order_no": "",
            "trace_no": "",
            "purchase_date": "2026-07-18",
        },
    )
    assert response.status_code == 200, response.text
    assert response.json()["purchase_order_no"] is None
    assert response.json()["trace_no"] is None
    assert response.json()["purchase_date"] == "2026-07-18"

    defaulted = await create_purchase_plan(client, headers, "默认申购单号", code="DQ-DEFAULT")
    response = await client.post(
        f"/api/v1/purchase-materials/{defaulted['id']}/move-to-record",
        headers=headers,
        json={"purchase_date": "2026-07-18"},
    )
    assert response.status_code == 200, response.text
    assert str(response.json()["purchase_order_no"]).startswith("申购 ")
    assert response.json()["trace_no"] is None


@pytest.mark.asyncio
async def test_purchase_excel_exports_use_json_template_specs(client: AsyncClient) -> None:
    headers = await auth_headers(client, "purchase")
    uncoded = await create_purchase_plan(client, headers, "待编码\u000b接触器")
    coded = await create_purchase_plan(
        client,
        headers,
        "已编码\u000c接触器",
        code="DQ-XLSX-1",
        actual_demand_person="不应导出的需求人",
        purchase_responsible="应导出的申购负责人",
    )

    code_export = await client.get(
        "/api/v1/purchase-materials/export-uncoded",
        headers=headers,
    )
    assert code_export.status_code == 200, code_export.text
    assert code_export.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    code_sheet = load_workbook(BytesIO(code_export.content)).active
    assert code_sheet["D7"].value == "待编码接触器"
    assert code_sheet["E7"].value == uncoded["model_spec"]
    assert code_sheet["I7"].value == uncoded["unit_name"]
    assert code_sheet["J7"].value == "否"
    assert code_sheet["L7"].value == "非资产"
    assert code_sheet["M7"].value == "当前无准确编码对应，需要新增编码"
    assert code_sheet["N7"].value == "HXNI 检修维护部"
    assert code_sheet.freeze_panes is None
    assert f"物料编码申请表_{date.today():%Y%m%d}.xlsx" in unquote(
        code_export.headers["content-disposition"]
    )

    purchase_export = await client.post(
        "/api/v1/purchase-materials/export-purchase-application",
        headers=headers,
        json={"material_ids": [coded["id"]]},
    )
    assert purchase_export.status_code == 200, purchase_export.text
    assert f"采购申请表_{date.today():%Y%m%d}.xlsx" in unquote(
        purchase_export.headers["content-disposition"]
    )
    purchase_sheet = load_workbook(BytesIO(purchase_export.content)).active
    assert purchase_sheet["A1"].value == "物料编码（必填）"
    assert purchase_sheet["A2"].value == coded["material_code"]
    assert purchase_sheet["B2"].value == "已编码接触器"
    assert str(purchase_sheet["C2"].value) == coded["planned_qty"]
    assert purchase_sheet["D2"].value == "应导出的申购负责人"
    assert purchase_sheet["E2"].value == "HXNI 检修维护部"
    assert purchase_sheet["G2"].value.date() == date.today() + timedelta(days=90)
    assert purchase_sheet["H2"].value == "正常"


@pytest.mark.asyncio
async def test_purchase_application_export_requires_code_subitem_and_usage(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    missing_code = await create_purchase_plan(client, headers, "缺编码计划")
    missing_subitem = await create_purchase_plan(
        client,
        headers,
        "缺子项号计划",
        code="DQ-XLSX-2",
        subitem_no=None,
    )
    missing_usage = await create_purchase_plan(
        client,
        headers,
        "缺用途计划",
        code="DQ-XLSX-3",
    )
    async with SessionLocal() as session:
        material = await session.get(PurchaseMaterial, int(missing_usage["id"]))
        assert material is not None
        material.usage = " "
        await session.commit()

    response = await client.post(
        "/api/v1/purchase-materials/export-purchase-application",
        headers=headers,
        json={
            "material_ids": [
                missing_code["id"],
                missing_subitem["id"],
                missing_usage["id"],
            ]
        },
    )

    assert response.status_code == 409, response.text
    payload = response.json()
    assert payload["code"] == "PURCHASE_APPLICATION_EXPORT_FIELDS_REQUIRED"
    assert payload["message"] == "导出采购申请表前请补全：编码、子项号、用途"
    assert payload["details"]["missing_fields"] == {
        "material_code": [missing_code["id"]],
        "subitem_no": [missing_subitem["id"]],
        "usage": [missing_usage["id"]],
    }


@pytest.mark.asyncio
async def test_purchase_result_exports_follow_filters_and_visible_columns(
    client: AsyncClient,
) -> None:
    headers = await auth_headers(client, "purchase")
    motor = await create_purchase_plan(client, headers, "导出电机", code="DQ-RESULT-1")
    await create_purchase_plan(client, headers, "不应导出的水泵", code="DQ-RESULT-2")

    plan_export = await client.post(
        "/api/v1/purchase-materials/export-results",
        headers=headers,
        json={
            "columns": ["name", "unit_name", "usage"],
            "name": "导出电机",
            "status": "正常",
        },
    )
    assert plan_export.status_code == 200, plan_export.text
    assert f"申购计划导出_{date.today():%Y%m%d}.xlsx" in unquote(
        plan_export.headers["content-disposition"]
    )
    plan_sheet = load_workbook(BytesIO(plan_export.content)).active
    assert [plan_sheet.cell(1, column).value for column in range(1, 4)] == [
        "名称",
        "计量单位",
        "用途",
    ]
    assert plan_sheet.max_column == 3
    assert plan_sheet.max_row == 2
    assert plan_sheet["A2"].value == motor["name"]
    assert plan_sheet["C2"].alignment.wrap_text is True

    record = await move_to_record(client, headers, int(motor["id"]))
    other_motor = await create_purchase_plan(
        client,
        headers,
        "导出电机",
        code="DQ-RESULT-3",
        actual_demand_person="其他需求人",
    )
    await move_to_record(client, headers, int(other_motor["id"]))
    record_export = await client.post(
        "/api/v1/purchase-records/export-results",
        headers=headers,
        json={
            "columns": ["material_name", "purchase_qty", "usage", "status"],
            "name": "导出电机",
            "actual_demand_person": record["actual_demand_person"],
            "status": "已申购",
        },
    )
    assert record_export.status_code == 200, record_export.text
    assert f"申购记录导出_{date.today():%Y%m%d}.xlsx" in unquote(
        record_export.headers["content-disposition"]
    )
    record_sheet = load_workbook(BytesIO(record_export.content)).active
    assert [record_sheet.cell(1, column).value for column in range(1, 5)] == [
        "物资",
        "申购数量",
        "用途",
        "状态",
    ]
    assert record_sheet.max_column == 4
    assert record_sheet.max_row == 2
    assert record["material_name"] in record_sheet["A2"].value
    assert record_sheet["B2"].value == f"{record['purchase_qty']} {record['unit_name']}"
    assert record_sheet["C2"].value == record["usage"]
    assert record_sheet["D2"].value == "已申购"


@pytest.mark.asyncio
async def test_missing_excel_template_returns_readable_400(
    client: AsyncClient, monkeypatch, tmp_path
) -> None:
    headers = await auth_headers(client, "purchase")
    monkeypatch.setattr(settings, "template_dir", tmp_path)

    response = await client.get("/api/v1/purchase-materials/export-uncoded", headers=headers)

    assert response.status_code == 400
    assert response.json()["code"] == "EXPORT_TEMPLATE_MISSING"
    assert "material-code-application.json" in response.json()["message"]
