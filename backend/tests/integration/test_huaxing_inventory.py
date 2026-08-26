from __future__ import annotations

import asyncio
import csv
from io import BytesIO, StringIO

import pytest
import xlwt
from httpx import AsyncClient
from openpyxl import Workbook

from tests.conftest import auth_headers

REPORT_HEADERS = [
    "首次入库日期",
    "货主",
    "仓库",
    "作业状态",
    "库位代码",
    "货品编码",
    "货品名称",
    "型号",
    "单位",
    "数量",
    "包装数量",
    "是否保税",
    "库存状态",
    "生产日期",
    "过期日期",
    "申购人",
    "申购部门",
    "追溯码",
    "子项号名称",
    "唛头号",
]


def build_report(rows: list[list[object]], *, first_header: str = "首次入库日期") -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    headers = list(REPORT_HEADERS)
    headers[0] = first_header
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def build_report_csv(rows: list[list[object]]) -> bytes:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(REPORT_HEADERS)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def build_report_xls(rows: list[list[object]]) -> bytes:
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Sheet1")
    for column, value in enumerate(REPORT_HEADERS):
        worksheet.write(0, column, value)
    for row_index, row in enumerate(rows, start=1):
        for column, value in enumerate(row):
            worksheet.write(row_index, column, value)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _row(
    material_code: str,
    name: str,
    *,
    warehouse: str = "P05综合仓",
    model_spec: str = "DN15",
    quantity: int = 25,
    unit: str = "个",
    purchaser: str = "吴冰",
    department: str = "生产调度中心",
    first_inbound_date: str = "2022-10-28",
    subitem: str = "201-冶炼主厂房",
) -> list[object]:
    return [
        first_inbound_date,
        "华星P05项目",
        warehouse,
        "正常",
        "HX-I2-05-00",
        material_code,
        name,
        model_spec,
        unit,
        quantity,
        quantity,
        "否",
        "良品",
        None,
        None,
        purchaser,
        department,
        "WJH00650076",
        subitem,
        "-",
    ]


async def _submit_import(
    client: AsyncClient, headers: dict[str, str], filename: str, content: bytes
):
    return await client.post(
        "/api/v1/huaxing-inventory/import",
        headers=headers,
        files={"file": (filename, content, "application/octet-stream")},
    )


async def _wait_job(client: AsyncClient, headers: dict[str, str], job_id: int) -> dict[str, object]:
    for _ in range(200):
        response = await client.get(
            f"/api/v1/huaxing-inventory/import-jobs/{job_id}", headers=headers
        )
        if response.status_code == 200:
            data = response.json()
            if data["status"] in ("SUCCEEDED", "FAILED"):
                return data
        await asyncio.sleep(0.05)
    raise AssertionError("import job did not reach terminal state in time")


async def _import_ok(
    client: AsyncClient, headers: dict[str, str], content: bytes, filename: str = "stock.xlsx"
) -> None:
    response = await _submit_import(client, headers, filename, content)
    assert response.status_code == 202, response.text
    job = await _wait_job(client, headers, response.json()["id"])
    assert job["status"] == "SUCCEEDED", job


@pytest.mark.asyncio
async def test_warehouse_import_and_search(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    await _import_ok(
        client,
        headers,
        build_report(
            [
                _row("L012-05048", "内丝三通", quantity=54506.04),
                _row(
                    "W004-00003",
                    "稀释剂",
                    warehouse="P06综合仓",
                    department="HXNI冶炼厂",
                    purchaser="夏军",
                    unit="桶",
                    quantity=3,
                ),
                _row("B001-00001", "接线端子", warehouse="P07综合仓"),
            ]
        ),
    )

    listed = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"page_size": 50}
    )
    assert listed.status_code == 200, listed.text
    assert listed.json()["total"] == 3

    by_name = await client.get(
        "/api/v1/huaxing-inventory",
        headers=headers,
        params={"name": "内丝三通"},
    )
    assert by_name.status_code == 200, by_name.text
    assert by_name.json()["total"] == 1
    item = by_name.json()["items"][0]
    assert item["material_code"] == "L012-05048"
    assert item["first_inbound_date"] == "2022-10-28"
    assert item["warehouse"] == "P05综合仓"
    assert item["name"] == "内丝三通"
    assert item["model_spec"] == "DN15"
    assert item["quantity"] == "54506.04"
    assert item["unit_name"] == "个"
    assert item["purchaser"] == "吴冰"
    assert item["purchase_department"] == "生产调度中心"
    assert item["subitem_no_name"] == "201-冶炼主厂房"

    by_warehouse = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"warehouse": "P06"}
    )
    assert by_warehouse.json()["total"] == 1
    assert by_warehouse.json()["items"][0]["material_code"] == "W004-00003"

    by_department = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"purchase_department": "冶炼厂"}
    )
    assert by_department.json()["total"] == 1

    no_match = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"name": "不存在"}
    )
    assert no_match.json()["items"] == []


@pytest.mark.asyncio
async def test_import_replaces_existing_rows(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    await _import_ok(
        client,
        headers,
        build_report([_row("Y001", "按钮"), _row("Y002", "接触器")]),
    )
    await _import_ok(
        client,
        headers,
        build_report([_row("Y999", "替换件")]),
    )
    listed = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"page_size": 50}
    )
    assert listed.json()["total"] == 1
    assert listed.json()["items"][0]["material_code"] == "Y999"


@pytest.mark.asyncio
async def test_header_alias_first_inbound_time(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    await _import_ok(
        client,
        headers,
        build_report([_row("Y001", "按钮")], first_header="首次入库时间"),
    )
    listed = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"material_code": "Y001"}
    )
    assert listed.json()["total"] == 1


@pytest.mark.asyncio
async def test_import_deduplicates_fully_identical_rows(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    dup = _row("Y001", "按钮")
    await _import_ok(
        client,
        headers,
        build_report([dup, dup, dup, _row("Y002", "接触器")]),
    )
    listed = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"page_size": 50}
    )
    assert listed.status_code == 200, listed.text
    assert listed.json()["total"] == 2
    # 完全相同行只保留一条；跨引用行数不受影响。
    by_code = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"material_code": "Y001"}
    )
    assert by_code.json()["total"] == 1


@pytest.mark.asyncio
async def test_missing_headers_marks_job_failed(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["编码", "名称"])  # 缺大部分必需列
    worksheet.append(["Y001", "按钮"])
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    response = await _submit_import(client, headers, "bad.xlsx", content.getvalue())
    assert response.status_code == 202, response.text
    job = await _wait_job(client, headers, response.json()["id"])
    assert job["status"] == "FAILED"
    assert job["error_code"] == "HUAXING_IMPORT_HEADERS_MISSING"


@pytest.mark.asyncio
async def test_import_permissions(client: AsyncClient) -> None:
    readonly_headers = await auth_headers(client, "readonly")
    purchase_headers = await auth_headers(client, "purchase")
    content = build_report([_row("Y001", "按钮")])

    listed = await client.get("/api/v1/huaxing-inventory", headers=readonly_headers)
    assert listed.status_code == 200

    readonly_import = await _submit_import(client, readonly_headers, "stock.xlsx", content)
    assert readonly_import.status_code == 403

    purchase_import = await _submit_import(client, purchase_headers, "stock.xlsx", content)
    assert purchase_import.status_code == 403


@pytest.mark.asyncio
async def test_import_rejects_unsupported_extension(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    response = await _submit_import(client, headers, "stock.txt", b"not excel")
    assert response.status_code == 400
    assert response.json()["code"] == "UNSUPPORTED_EXCEL_FILE"


@pytest.mark.asyncio
async def test_import_csv(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    response = await _submit_import(
        client, headers, "stock.csv", build_report_csv([_row("C001", "按钮")])
    )
    assert response.status_code == 202, response.text
    job = await _wait_job(client, headers, response.json()["id"])
    assert job["status"] == "SUCCEEDED", job
    assert job["result"]["imported_count"] == 1

    listed = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"material_code": "C001"}
    )
    assert listed.status_code == 200, listed.text
    assert listed.json()["total"] == 1
    assert listed.json()["items"][0]["material_code"] == "C001"


@pytest.mark.asyncio
async def test_import_xls(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    response = await _submit_import(
        client, headers, "stock.xls", build_report_xls([_row("X001", "接触器")])
    )
    assert response.status_code == 202, response.text
    job = await _wait_job(client, headers, response.json()["id"])
    assert job["status"] == "SUCCEEDED", job
    assert job["result"]["imported_count"] == 1

    listed = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"material_code": "X001"}
    )
    assert listed.status_code == 200, listed.text
    assert listed.json()["items"][0]["material_code"] == "X001"


@pytest.mark.asyncio
async def test_filter_by_model_spec(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    await _import_ok(
        client,
        headers,
        build_report(
            [
                _row("M001", "接触器", model_spec="CJX2-2510"),
                _row("M002", "接触器", model_spec="CJX2-1210"),
            ]
        ),
    )
    listed = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"model_spec": "CJX2-25"}
    )
    assert listed.status_code == 200, listed.text
    assert listed.json()["total"] == 1
    assert listed.json()["items"][0]["material_code"] == "M001"


@pytest.mark.asyncio
async def test_filter_code_or_and_cross_field_and(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    await _import_ok(
        client,
        headers,
        build_report(
            [
                _row("L012-05048", "内丝三通", model_spec="DN15"),
                _row("W004-00003", "稀释剂", model_spec="25L", quantity=3),
                _row("L013-05048", "内丝弯头", model_spec="DN15"),
            ]
        ),
    )
    # 同一输入框内用 | 分隔多关键词按 OR 匹配：命中两条含 L012/L013 的编码。
    by_code_or = await client.get(
        "/api/v1/huaxing-inventory",
        headers=headers,
        params={"material_code": "L012|L013"},
    )
    assert by_code_or.status_code == 200, by_code_or.text
    assert by_code_or.json()["total"] == 2

    # 名称框内 OR：命中 内丝三通/内丝弯头 两条。
    by_name_or = await client.get(
        "/api/v1/huaxing-inventory", headers=headers, params={"name": "内丝三通|内丝弯头"}
    )
    assert by_name_or.json()["total"] == 2

    # 不同输入框之间按 AND：编码 L012 且名称 内丝三通 → 仅一条。
    by_and = await client.get(
        "/api/v1/huaxing-inventory",
        headers=headers,
        params={"material_code": "L012", "name": "内丝三通"},
    )
    assert by_and.status_code == 200, by_and.text
    assert by_and.json()["total"] == 1
    assert by_and.json()["items"][0]["material_code"] == "L012-05048"
