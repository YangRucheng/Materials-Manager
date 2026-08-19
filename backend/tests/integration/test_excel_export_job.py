from __future__ import annotations

import asyncio
import time
from datetime import timedelta
from pathlib import Path
from urllib.parse import unquote

import pytest
from httpx import AsyncClient
from sqlalchemy import select

import app.services.purchase_request_service as record_service
from app.core.constants import EXPORT_ROW_LIMIT
from app.core.database import SessionLocal
from app.core.errors import AppError
from app.domain.enums import ExcelExportJobStatus
from app.models import ExcelExportJob, PurchaseMaterial
from app.services import excel_export_job_service
from app.services.common import utcnow
from tests.conftest import auth_headers, await_export_job


async def _insert_job(
    *,
    export_type: str,
    status: ExcelExportJobStatus,
    file_path: str | None = None,
    created_by: int | None = None,
    finished_at=None,
) -> int:
    async with SessionLocal() as session:
        job = ExcelExportJob(
            export_type=export_type,
            status=status,
            file_path=file_path,
            created_by=created_by,
            finished_at=finished_at,
        )
        session.add(job)
        await session.commit()
        await session.refresh(job)
        return job.id


@pytest.mark.asyncio
async def test_mark_stale_exports_failed_marks_interrupted_and_unlinks(
    client: AsyncClient, tmp_path
) -> None:
    paths = [tmp_path / f"e{i}.xlsx" for i in range(4)]
    for path in paths:
        await asyncio.to_thread(path.write_bytes, b"x")
    statuses = [
        ExcelExportJobStatus.PENDING,
        ExcelExportJobStatus.RUNNING,
        ExcelExportJobStatus.SUCCEEDED,
        ExcelExportJobStatus.FAILED,
    ]
    async with SessionLocal() as session:
        for status, path in zip(statuses, paths, strict=False):
            session.add(
                ExcelExportJob(
                    export_type="STALE_TEST",
                    status=status,
                    file_path=str(path),
                )
            )
        await session.commit()

    count = await excel_export_job_service.mark_stale_exports_failed()

    assert count == 2
    async with SessionLocal() as session:
        rows = list(
            (await session.scalars(select(ExcelExportJob).order_by(ExcelExportJob.id))).all()
        )
    assert [row.status for row in rows] == [
        ExcelExportJobStatus.FAILED,
        ExcelExportJobStatus.FAILED,
        ExcelExportJobStatus.SUCCEEDED,
        ExcelExportJobStatus.FAILED,
    ]
    assert rows[0].error_code == "SERVER_RESTARTED"
    assert rows[1].error_code == "SERVER_RESTARTED"
    assert rows[2].error_code is None
    assert rows[3].error_code is None
    for index, path in enumerate(paths):
        assert path.exists() == (index >= 2)


@pytest.mark.asyncio
async def test_cleanup_finished_exports_deletes_only_old_terminal_rows_and_files(
    client: AsyncClient, tmp_path
) -> None:
    now = utcnow()
    old_path = tmp_path / "old.xlsx"
    old_path.write_bytes(b"x")
    recent_path = tmp_path / "recent.xlsx"
    recent_path.write_bytes(b"x")
    async with SessionLocal() as session:
        old_finished = ExcelExportJob(
            export_type="CLEANUP_TEST",
            status=ExcelExportJobStatus.SUCCEEDED,
            file_path=str(old_path),
            finished_at=now - timedelta(days=60),
        )
        recent_finished = ExcelExportJob(
            export_type="CLEANUP_TEST",
            status=ExcelExportJobStatus.FAILED,
            file_path=str(recent_path),
            finished_at=now - timedelta(days=1),
        )
        running = ExcelExportJob(
            export_type="CLEANUP_TEST",
            status=ExcelExportJobStatus.RUNNING,
            file_path=None,
        )
        session.add_all([old_finished, recent_finished, running])
        await session.commit()

    count = await excel_export_job_service.cleanup_finished_exports(retention_days=30)

    assert count == 1
    assert not old_path.exists()
    assert recent_path.exists()
    async with SessionLocal() as session:
        remaining = list(
            (
                await session.scalars(select(ExcelExportJob).order_by(ExcelExportJob.id))
            ).all()
        )
    assert [row.export_type for row in remaining] == ["CLEANUP_TEST", "CLEANUP_TEST"]
    assert [row.status for row in remaining] == [
        ExcelExportJobStatus.FAILED,
        ExcelExportJobStatus.RUNNING,
    ]


@pytest.mark.asyncio
async def test_run_job_success_keeps_file_and_marks_succeeded(
    client: AsyncClient, tmp_path
) -> None:
    target = tmp_path / "job.xlsx"

    async def ok_processor(path: Path) -> dict[str, object]:
        assert path == target
        await asyncio.to_thread(path.write_bytes, b"xlsx-bytes")
        return {"download_filename": "申购计划导出.xlsx", "rows": 3, "image_count": 2}

    job_id = await _insert_job(
        export_type="UNIT", status=ExcelExportJobStatus.PENDING, file_path=str(target)
    )
    await excel_export_job_service._run_job(job_id, ok_processor)

    async with SessionLocal() as session:
        job = await session.get(ExcelExportJob, job_id)
    assert job is not None
    assert job.status == ExcelExportJobStatus.SUCCEEDED
    assert job.download_filename == "申购计划导出.xlsx"
    assert job.result == {"download_filename": "申购计划导出.xlsx", "rows": 3, "image_count": 2}
    assert job.error_code is None
    assert target.exists()


@pytest.mark.asyncio
async def test_run_job_app_error_marks_failed_and_unlinks_partial_file(
    client: AsyncClient, tmp_path
) -> None:
    target = tmp_path / "job.xlsx"

    async def failing_processor(path: Path) -> dict[str, object]:
        await asyncio.to_thread(path.write_bytes, b"partial")
        raise AppError("EXPORT_RESULT_LIMIT_EXCEEDED", "查询结果超过上限")

    job_id = await _insert_job(
        export_type="UNIT", status=ExcelExportJobStatus.PENDING, file_path=str(target)
    )
    await excel_export_job_service._run_job(job_id, failing_processor)

    async with SessionLocal() as session:
        job = await session.get(ExcelExportJob, job_id)
    assert job is not None
    assert job.status == ExcelExportJobStatus.FAILED
    assert job.error_code == "EXPORT_RESULT_LIMIT_EXCEEDED"
    assert job.error_message == "查询结果超过上限"
    assert not target.exists()


@pytest.mark.asyncio
async def test_run_job_invalid_result_marks_failed(client: AsyncClient, tmp_path) -> None:
    target = tmp_path / "job.xlsx"

    async def empty_processor(path: Path) -> dict[str, object]:
        await asyncio.to_thread(path.write_bytes, b"x")
        return {"rows": 1}  # 缺少 download_filename

    job_id = await _insert_job(
        export_type="UNIT", status=ExcelExportJobStatus.PENDING, file_path=str(target)
    )
    await excel_export_job_service._run_job(job_id, empty_processor)

    async with SessionLocal() as session:
        job = await session.get(ExcelExportJob, job_id)
    assert job is not None
    assert job.status == ExcelExportJobStatus.FAILED
    assert job.error_code == "INTERNAL_EXPORT_ERROR"
    assert not target.exists()


@pytest.mark.asyncio
async def test_run_job_unexpected_error_marks_failed_and_unlinks(
    client: AsyncClient, tmp_path
) -> None:
    target = tmp_path / "job.xlsx"

    async def crashing_processor(path: Path) -> dict[str, object]:
        await asyncio.to_thread(path.write_bytes, b"partial")
        raise RuntimeError("boom")

    job_id = await _insert_job(
        export_type="UNIT", status=ExcelExportJobStatus.PENDING, file_path=str(target)
    )
    await excel_export_job_service._run_job(job_id, crashing_processor)

    async with SessionLocal() as session:
        job = await session.get(ExcelExportJob, job_id)
    assert job is not None
    assert job.status == ExcelExportJobStatus.FAILED
    assert job.error_code == "INTERNAL_EXPORT_ERROR"
    assert job.error_message == "导出任务发生未知错误"
    assert not target.exists()


@pytest.mark.asyncio
async def test_enqueue_export_returns_pending_job_and_job_runs_to_completion(
    client: AsyncClient, tmp_path
) -> None:
    async def processor(path: Path) -> dict[str, object]:
        await asyncio.to_thread(path.write_bytes, b"xlsx")
        return {"download_filename": "enqueued.xlsx", "rows": 1}

    job = await excel_export_job_service.enqueue_export(
        export_type="UNIT",
        params={"columns": ["name"]},
        processor=processor,
        created_by=None,
    )
    assert job.status == ExcelExportJobStatus.PENDING
    assert job.export_type == "UNIT"
    assert job.params == {"columns": ["name"]}

    deadline = time.monotonic() + 2.0
    while True:
        async with SessionLocal() as session:
            row = await session.get(ExcelExportJob, job.id)
        if row.status == ExcelExportJobStatus.SUCCEEDED:
            break
        assert time.monotonic() < deadline, row.status
        await asyncio.sleep(0.01)


@pytest.mark.asyncio
async def test_record_export_async_flow_with_images(client: AsyncClient, tmp_path) -> None:
    from io import BytesIO

    from openpyxl import load_workbook
    from PIL import Image

    from tests.integration.test_procurement import create_purchase_plan, move_to_record

    headers = await auth_headers(client, "purchase")
    source = BytesIO()
    Image.new("RGB", (64, 40), "orange").save(source, format="PNG")
    uploaded = await client.post(
        "/api/v1/files/images",
        headers=headers,
        files={"file": ("export.png", source.getvalue(), "image/png")},
    )
    assert uploaded.status_code == 201, uploaded.text
    file_id = uploaded.json()["id"]
    plan = await create_purchase_plan(
        client, headers, "异步导出电机", code="ASYNC-EXP-1", image_ids=[file_id]
    )
    await move_to_record(client, headers, int(plan["id"]))

    created = await client.post(
        "/api/v1/purchase-records/export-results",
        headers=headers,
        json={"columns": ["material_name", "images"], "name": "异步导出电机"},
    )
    assert created.status_code == 202, created.text
    job = await await_export_job(client, headers, created.json()["id"])
    assert job["status"] == "SUCCEEDED"
    assert job["download_filename"].startswith("申购记录导出_")
    assert job["download_filename"].endswith(".xlsx")
    assert job["file_uuid"]
    assert job["result"]["rows"] == 1
    assert job["result"]["image_count"] == 1

    # 下载端点不鉴权：不带 Authorization 头直接按 file_uuid 拉取。
    download = await client.get(f"/api/v1/excel-export-jobs/files/{job['file_uuid']}")
    assert download.status_code == 200, download.text
    assert "申购记录导出_" in unquote(download.headers["content-disposition"])
    sheet = load_workbook(BytesIO(download.content)).active
    assert len(sheet._images) == 1
    assert sheet["A2"].value == plan["name"]


@pytest.mark.asyncio
async def test_plan_export_async_flow_and_ownership(client: AsyncClient, tmp_path) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from tests.integration.test_procurement import create_purchase_plan

    headers = await auth_headers(client, "purchase")
    other_headers = await auth_headers(client, "warehouse")
    await create_purchase_plan(client, headers, "异步计划电机", code="ASYNC-PLN-1")

    created = await client.post(
        "/api/v1/purchase-materials/export-results",
        headers=headers,
        json={"columns": ["name", "usage"], "name": "异步计划电机"},
    )
    assert created.status_code == 202, created.text
    job = await await_export_job(client, headers, created.json()["id"])
    assert job["status"] == "SUCCEEDED"
    assert job["file_uuid"]

    # 状态轮询仍鉴权 + 属主校验：他人不可见。
    foreign = await client.get(f"/api/v1/excel-export-jobs/{job['id']}", headers=other_headers)
    assert foreign.status_code == 400
    assert foreign.json()["code"] == "NOT_FOUND"

    # 下载端点不鉴权（凭 uuid7 匿名拉取）：无 Authorization 头也可下载。
    download = await client.get(f"/api/v1/excel-export-jobs/files/{job['file_uuid']}")
    assert download.status_code == 200, download.text
    sheet = load_workbook(BytesIO(download.content)).active
    assert sheet["A2"].value == "异步计划电机"


@pytest.mark.asyncio
async def test_download_guards_not_ready_and_expired(
    client: AsyncClient, tmp_path
) -> None:
    from tests.integration.test_procurement import create_purchase_plan

    headers = await auth_headers(client, "purchase")

    # 不存在的 uuid 下载 → EXPORT_FILE_EXPIRED
    missing_uuid = "00000000-0000-7000-8000-000000000000"
    pending_download = await client.get(
        f"/api/v1/excel-export-jobs/files/{missing_uuid}"
    )
    assert pending_download.status_code == 400
    assert pending_download.json()["code"] == "EXPORT_FILE_EXPIRED"

    # 成功任务删文件后下载 → EXPORT_FILE_EXPIRED（他人查状态仍 NOT_FOUND）
    other_headers = await auth_headers(client, "warehouse")
    await create_purchase_plan(client, headers, "过期电机", code="ASYNC-EXPIRED")
    created = await client.post(
        "/api/v1/purchase-materials/export-results",
        headers=headers,
        json={"columns": ["name"], "name": "过期电机"},
    )
    assert created.status_code == 202, created.text
    job = await await_export_job(client, headers, created.json()["id"])
    async with SessionLocal() as session:
        row = await session.get(ExcelExportJob, job["id"])
        target = Path(row.file_path)
        await asyncio.to_thread(target.unlink)
    expired = await client.get(
        f"/api/v1/excel-export-jobs/files/{job['file_uuid']}"
    )
    assert expired.status_code == 400
    assert expired.json()["code"] == "EXPORT_FILE_EXPIRED"

    missing = await client.get("/api/v1/excel-export-jobs/999999", headers=headers)
    assert missing.status_code == 400
    assert missing.json()["code"] == "NOT_FOUND"
    missing_file = await client.get(
        f"/api/v1/excel-export-jobs/files/{missing_uuid}", headers=other_headers
    )
    assert missing_file.status_code == 400
    assert missing_file.json()["code"] == "EXPORT_FILE_EXPIRED"


@pytest.mark.asyncio
async def test_export_limit_exceeded_surfaces_as_job_failure(
    client: AsyncClient, monkeypatch
) -> None:
    from tests.integration.test_procurement import create_purchase_plan

    headers = await auth_headers(client, "purchase")
    await create_purchase_plan(client, headers, "超限电机", code="ASYNC-BIG")

    async def fake_search(session, **kwargs) -> tuple[list[PurchaseMaterial], int]:
        return [], EXPORT_ROW_LIMIT + 1

    monkeypatch.setattr(record_service, "search_purchase_records", fake_search)

    created = await client.post(
        "/api/v1/purchase-records/export-results",
        headers=headers,
        json={"columns": ["material_name"], "name": "超限电机"},
    )
    assert created.status_code == 202, created.text
    job = await await_export_job(client, headers, created.json()["id"])
    assert job["status"] == "FAILED"
    assert job["error_code"] == "EXPORT_RESULT_LIMIT_EXCEEDED"
    assert "10000" in (job["error_message"] or "")