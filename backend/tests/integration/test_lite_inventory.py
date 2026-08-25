from __future__ import annotations

import asyncio
import csv
from io import BytesIO, StringIO

import pytest
import xlwt
from httpx import AsyncClient
from openpyxl import Workbook

from tests.conftest import auth_headers

LITE_HEADERS = ["物资名称", "型号规格", "单位", "数量", "备注"]


def build_report(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(LITE_HEADERS)
    for row in rows:
        worksheet.append(row)
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def build_report_csv(rows: list[list[object]]) -> bytes:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(LITE_HEADERS)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def build_report_xls(rows: list[list[object]]) -> bytes:
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Sheet1")
    for column, value in enumerate(LITE_HEADERS):
        worksheet.write(0, column, value)
    for row_index, row in enumerate(rows, start=1):
        for column, value in enumerate(row):
            worksheet.write(row_index, column, value)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _row(
    name: str,
    *,
    model_spec: str = "CJX2-2510",
    unit: str = "个",
    quantity: int = 10,
    remark: str = "",
) -> list[object]:
    return [name, model_spec, unit, quantity, remark]


async def _submit_import(
    client: AsyncClient, headers: dict[str, str], filename: str, content: bytes
):
    return await client.post(
        "/api/v1/secondary-warehouse/import",
        headers=headers,
        files={"file": (filename, content, "application/octet-stream")},
    )


async def _wait_job(client: AsyncClient, headers: dict[str, str], job_id: int) -> dict[str, object]:
    for _ in range(200):
        response = await client.get(
            f"/api/v1/secondary-warehouse/import-jobs/{job_id}", headers=headers
        )
        if response.status_code == 200:
            data = response.json()
            if data["status"] in ("SUCCEEDED", "FAILED"):
                return data
        await asyncio.sleep(0.05)
    raise AssertionError("import job did not reach terminal state in time")


async def _import_ok(
    client: AsyncClient, headers: dict[str, str], content: bytes, filename: str = "lite.xlsx"
) -> dict[str, object]:
    response = await _submit_import(client, headers, filename, content)
    assert response.status_code == 202, response.text
    job = await _wait_job(client, headers, response.json()["id"])
    assert job["status"] == "SUCCEEDED", job
    return job


@pytest.mark.asyncio
async def test_warehouse_import_and_search(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    job = await _import_ok(
        client,
        headers,
        build_report(
            [
                _row("交流接触器", quantity=5, remark="库存"),
                _row("热继电器", model_spec="JRS1-25", unit="只", quantity=12),
            ]
        ),
    )
    assert job["result"]["imported_count"] == 2

    listed = await client.get(
        "/api/v1/secondary-warehouse", headers=headers, params={"page_size": 50}
    )
    assert listed.status_code == 200, listed.text
    assert listed.json()["total"] == 2

    by_keyword = await client.get(
        "/api/v1/secondary-warehouse",
        headers=headers,
        params={"keyword": "交流接触器"},
    )
    assert by_keyword.status_code == 200, by_keyword.text
    assert by_keyword.json()["total"] == 1
    item = by_keyword.json()["items"][0]
    assert item["name"] == "交流接触器"
    assert item["model_spec"] == "CJX2-2510"
    assert item["unit_name"] == "个"
    assert item["quantity"] == "5"
    assert item["remark"] == "库存"

    no_match = await client.get(
        "/api/v1/secondary-warehouse", headers=headers, params={"keyword": "不存在"}
    )
    assert no_match.json()["items"] == []


@pytest.mark.asyncio
async def test_import_replaces_existing_rows(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    await _import_ok(client, headers, build_report([_row("按钮"), _row("接触器")]))
    await _import_ok(client, headers, build_report([_row("替换件")]))
    listed = await client.get(
        "/api/v1/secondary-warehouse", headers=headers, params={"page_size": 50}
    )
    assert listed.json()["total"] == 1
    assert listed.json()["items"][0]["name"] == "替换件"


@pytest.mark.asyncio
async def test_import_deduplicates_fully_identical_rows(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    dup = _row("按钮")
    await _import_ok(client, headers, build_report([dup, dup, dup, _row("接触器")]))
    listed = await client.get(
        "/api/v1/secondary-warehouse", headers=headers, params={"page_size": 50}
    )
    assert listed.status_code == 200, listed.text
    assert listed.json()["total"] == 2


@pytest.mark.asyncio
async def test_missing_headers_marks_job_failed(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["编码", "名称"])
    worksheet.append(["Y001", "按钮"])
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    response = await _submit_import(client, headers, "bad.xlsx", content.getvalue())
    assert response.status_code == 202, response.text
    job = await _wait_job(client, headers, response.json()["id"])
    assert job["status"] == "FAILED"
    assert job["error_code"] == "LITE_IMPORT_HEADERS_MISSING"


@pytest.mark.asyncio
async def test_missing_name_marks_job_failed(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(LITE_HEADERS)
    worksheet.append(["", "CJX2-2510", "个", 3, ""])
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    response = await _submit_import(client, headers, "bad.xlsx", content.getvalue())
    assert response.status_code == 202, response.text
    job = await _wait_job(client, headers, response.json()["id"])
    assert job["status"] == "FAILED"
    assert job["error_code"] == "LITE_IMPORT_NAME_REQUIRED"


@pytest.mark.asyncio
async def test_invalid_quantity_marks_job_failed(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(LITE_HEADERS)
    worksheet.append(["接触器", "CJX2-2510", "个", "not-a-number", ""])
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    response = await _submit_import(client, headers, "bad.xlsx", content.getvalue())
    assert response.status_code == 202, response.text
    job = await _wait_job(client, headers, response.json()["id"])
    assert job["status"] == "FAILED"
    assert job["error_code"] == "LITE_IMPORT_INVALID_QUANTITY"


@pytest.mark.asyncio
async def test_import_permissions(client: AsyncClient) -> None:
    readonly_headers = await auth_headers(client, "readonly")
    purchase_headers = await auth_headers(client, "purchase")
    content = build_report([_row("按钮")])

    listed = await client.get("/api/v1/secondary-warehouse", headers=readonly_headers)
    assert listed.status_code == 200

    readonly_import = await _submit_import(client, readonly_headers, "lite.xlsx", content)
    assert readonly_import.status_code == 403

    purchase_import = await _submit_import(client, purchase_headers, "lite.xlsx", content)
    assert purchase_import.status_code == 403


@pytest.mark.asyncio
async def test_import_rejects_unsupported_extension(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    response = await _submit_import(client, headers, "lite.txt", b"not excel")
    assert response.status_code == 400
    assert response.json()["code"] == "UNSUPPORTED_EXCEL_FILE"


@pytest.mark.asyncio
async def test_import_csv_and_xls(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    job = await _import_ok(client, headers, build_report_csv([_row("按钮")]), "lite.csv")
    assert job["result"]["imported_count"] == 1
    listed = await client.get(
        "/api/v1/secondary-warehouse", headers=headers, params={"keyword": "按钮"}
    )
    assert listed.json()["total"] == 1

    job_xls = await _import_ok(client, headers, build_report_xls([_row("接触器")]), "lite.xls")
    assert job_xls["result"]["imported_count"] == 1
    listed = await client.get(
        "/api/v1/secondary-warehouse", headers=headers, params={"keyword": "接触器"}
    )
    assert listed.json()["total"] == 1


@pytest.mark.asyncio
async def test_last_import(client: AsyncClient) -> None:
    headers = await auth_headers(client, "warehouse")
    before = await client.get("/api/v1/secondary-warehouse/last-import", headers=headers)
    assert before.status_code == 200
    assert before.json()["last_import_at"] is None

    await _import_ok(client, headers, build_report([_row("按钮")]))
    after = await client.get("/api/v1/secondary-warehouse/last-import", headers=headers)
    assert after.status_code == 200
    assert after.json()["last_import_at"] is not None
