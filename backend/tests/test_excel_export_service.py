from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.core.config import Settings, settings
from app.core.errors import AppError
from app.services import excel_export_service


def test_default_template_directory_is_relative_to_working_directory(monkeypatch) -> None:
    monkeypatch.delenv("APP_TEMPLATE_DIR", raising=False)

    configured = Settings(_env_file=None)  # type: ignore[call-arg]

    assert configured.template_dir == Path.cwd() / "data" / "template"


def test_load_spec_reads_configured_runtime_template(monkeypatch) -> None:
    source_template_dir = Path(__file__).parents[2] / "example" / "template"
    monkeypatch.setattr(settings, "template_dir", source_template_dir)

    spec = excel_export_service._load_spec("purchase-application.json")

    assert spec["template_name"] == "采购申请模板"


def test_missing_template_raises_readable_service_error(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(settings, "template_dir", tmp_path)

    with pytest.raises(AppError) as raised:
        excel_export_service._load_spec("purchase-application.json")

    assert raised.value.status_code == 400
    assert raised.value.code == "EXPORT_TEMPLATE_MISSING"
    assert "purchase-application.json" in raised.value.message


def test_invalid_template_raises_readable_business_error(monkeypatch, tmp_path: Path) -> None:
    (tmp_path / "purchase-application.json").write_text("{}", encoding="utf-8")
    monkeypatch.setattr(settings, "template_dir", tmp_path)

    with pytest.raises(AppError) as raised:
        excel_export_service.render_excel("purchase-application.json", [])

    assert raised.value.status_code == 400
    assert raised.value.code == "EXPORT_TEMPLATE_INVALID"


def test_result_excel_uses_visible_columns_and_readable_layout() -> None:
    content, filename = excel_export_service.render_result_excel(
        "申购计划导出",
        [("name", "名称"), ("usage", "用途")],
        [{"name": "电机", "usage": "控制柜检修\n备用设备维护"}],
    )

    assert filename == f"申购计划导出_{date.today():%Y%m%d}.xlsx"
    sheet = load_workbook(BytesIO(content)).active
    assert [sheet.cell(1, column).value for column in range(1, 3)] == ["名称", "用途"]
    assert sheet["A2"].value == "电机"
    assert sheet["B2"].value == "控制柜检修\n备用设备维护"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:B2"
    assert sheet.row_dimensions[1].height == 32
    assert sheet.row_dimensions[2].height == 30
    assert sheet["B2"].alignment.wrap_text is True
    assert sheet.column_dimensions["A"].width >= 14
    assert sheet.column_dimensions["B"].width >= 14
